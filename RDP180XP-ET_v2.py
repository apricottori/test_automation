# #############################
# RDP180XP REV0 Electrical Test
# Date : 2024/11/23
# Author : Ron
# Program name : RDP180XP-ETV2
###############################

import time
import inspect
import tkinter
import tkinter.ttk
import openpyxl as op
from datetime import datetime
from raonpy.device.keithley2401 import KEITHLEY2401
from raonpy.device.agilent34401a import Agilent34401A # For DMM
from raonpy.device.e36312a import E36312A # For Power supply 
import raonpy.device.su241 as su # For Chamber
from raonpy.rti.efm8evb import EVB


def update_gui_info(num_item):
    global func_name
    global temp_curr
    global cnt_item

    temp_curr = chamber.get_current_temp()
    cnt_item += 1
    new_func_name = ' '.join(func_name.split("_")[1:])
    var.set(((total_progress / (total_item_cal + (total_item_noncal * (len(temp_lst)-1)))) * 100))
    label.config(text=f'Progress Rate : {((total_progress / (total_item_cal + (total_item_noncal * (len(temp_lst)-1)))) * 100):.0f}%')
    label2.config(text=f'Completed Test Item : {new_func_name} ({cnt_item}/{num_item})')
    label3.config(text=f'Current Test Temperature : {temp_curr:.1f}°C ({temp_progress+1}/{len(temp_lst)})')
    root.update()

def option_create_sheet():
    if logfile_path == None:
        wb = op.Workbook()
        ws = wb.active
        ws.title = f"#{chip_num}"
    elif logfile_path != None:
        wb = op.load_workbook(logfile_path)
        ws_names = wb.sheetnames
        for i in ws_names:
            if chip_num == i[1:]:
                ws = wb[f"#{chip_num}"]
                break
            elif chip_num != i[1:]:
                wb.create_sheet(f"#{chip_num}")
                ws = wb[f"#{chip_num}"]
                test = wb.sheetnames
                if test[-1] != f"#{chip_num}":
                    del wb[test[-1]]
                break
    return wb, ws

def record_test_info():
    print("RECORD TEST INFO")
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    wb, ws = option_create_sheet()

    global row_usl
    global row_lsl
    row_usl = 10
    row_lsl = 11

    ws.cell(row=1, column=1, value=f"Tester Name : {tester_name}")
    ws.cell(row=2, column=1, value="Measuring equipment : KEITHLEY2410")
    ws.cell(row=3, column=1, value="Program Name : RDP180XP-ETV0")
    ws.cell(row=4, column=1, value="Lot number : N/A")
    ws.cell(row=5, column=1, value=f"EVBoard number : {evboard_num}")
    ws.cell(row=6, column=1, value=f"Start DATE : {now.date()}")
    ws.cell(row=7, column=1, value=f"Start TIME : {time.strftime('%X')}")
    ws.cell(row=9, column=5, value="UNIT")
    ws.cell(row=10, column=5, value="USL")
    ws.cell(row=11, column=5, value="LSL")
    # Test item write
    ws.cell(row=8, column=col2, value="Test_Addr") # From register map
    ws.cell(row=8, column=col2+1, value="Init_Value") # From register map
    ws.cell(row=8, column=col2+2, value="Init_Read_Value") # Real i2c read data
    ws.cell(row=8, column=col2+3, value="Init_Read_Result") # Init value check
    ws.cell(row=8, column=col2+4, value="Asc_All_0x55_Read_Value") # 0x55 All Write(asc)
    ws.cell(row=8, column=col2+5, value="Asc_All_Read_Result") # 0x55 Read(P/F)
    ws.cell(row=8, column=col2+6, value="Asc_Read0_0xAA_Value") # Asc 0xaa Write
    ws.cell(row=8, column=col2+7, value="Asc_Read0_Result") # Asc 0xaa Read(P/F)
    ws.cell(row=8, column=col2+8, value="AsctoDesc_Read_0x55_Value") # Asc 0x55 Write
    ws.cell(row=8, column=col2+9, value="AsctoDesc_Read_Result") # Asc_Desc 0x55 Read(P/F)
    ws.cell(row=8, column=col2+10, value="Desc_Read0_0xAA_Value") # Desc 0xaa Write
    ws.cell(row=8, column=col2+11, value="Desc_Read0_Result") # Desc 0xaa Read(P/F)
    ws.cell(row=8, column=col2+12, value="Desc_Read1_0x55_Value") # Desc 0x55 Write
    ws.cell(row=8, column=col2+13, value="Desc_Read1_Result") # Desc 0x55 Read(P/F)

    wb.save(save_path)
    wb.close()

def i2c_reset():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x04)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x00)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x02)
    time.sleep(1/2000) 
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x00)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x01)
    time.sleep(1/2000) 
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0004, 0x00)
    time.sleep(1/2000) 
    print("I2C_RESET Done")

def i2c_all_power_up():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    evb.i2c0_reg16_write(SlaveAddr_0, 0x0005, 0x00)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0006, 0x00)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0007, 0x00)
    time.sleep(1/2000)
    print("I2C_Power-up Done")

def i2c_power_up_ts():
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0005, 0x00)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0006, 0xFF)
    time.sleep(1/2000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0007, 0xF0)
    time.sleep(1/2000)
    print("I2C_Power-up_ts Done")

def i2c_all_register_test():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----I2C Function Test START-----")
    wb = op.load_workbook(regmap_path)
    # ws = wb.active
    ws = wb["RegMap"]
    with open(regmap_path, 'r', encoding='utf-8') as file:
        exc_list = [3, 4, 5, 6, 7, 15, 1063] # exclusion address

        if logfile_path == None:
            wb2 = op.Workbook()
            ws2 = wb2.active
            ws2.title = f"#{chip_num}"
        elif logfile_path != None:
            wb2 = op.load_workbook(logfile_path)
            ws2_names = wb2.sheetnames
            for i in ws2_names:
                if chip_num == i[1:]:
                    ws2 = wb2[f"#{chip_num}"]
                    break
                elif chip_num != i[1:]:
                    wb2.create_sheet(f"#{chip_num}")
                    ws2 = wb2[f"#{chip_num}"]
                    test = wb2.sheetnames
                    if test[-1] != f"#{chip_num}":
                        del wb2[test[-1]]
                    break

        mode = ws["B"]
        row = 6 # Reading register map file row
        row2 = 12 # log file start row
        cnt = 0
        ##### Init Read #####
        print("-----ALL REGISTER RESET-----")
        i2c_reset()
        print("-----INITIAL VALUE READ START-----")
        for idx in range(row, len(mode)+1):
            data = ws.cell(row=idx, column=2).value
            init_val = ws.cell(row=idx, column=3).value

            if data[:2] == "RW" and init_val != "UNUSED":
                Addr = int(data[2:], 16)
                read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr))
                
                ws2.cell(row=row2, column=col2, value=data)
                ws2.cell(row=row2, column=col2+1, value=init_val)
                ws2.cell(row=row2, column=col2+2, value=read_data)

                if int(init_val, 16) == int(read_data, 16):
                    ws2.cell(row=row2, column=col2+3, value="PASS")
                    # ws2.cell(row=row2, column=9).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                else:
                    ws2.cell(row=row2, column=col2+3, value="FAIL")
                    # ws2.cell(row=row2, column=9).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                    cnt += 1

                row2 += 1
                continue
            elif data[:2] == "RO":                
                continue
            elif data[:2] == "WO":               
                continue
            elif data == None:
                break
        
        mode_des = ws2["A"]
        ws2.cell(row=len(mode_des)+2, column=col2, value="FAIL Count")
        ws2.cell(row=len(mode_des)+2, column=col2+3, value=cnt)
        
        ##### Ascending All Write 0x55 #####
        print("-----ASCENDING WRITE_0x55 START-----")
        cnt2 = 0
        write_data0 = 0x55
        for idx2 in range(row, len(mode)+1):

            data = ws.cell(row=idx2, column=2).value
            init_val = ws.cell(row=idx2, column=3).value
            reg_name = ws.cell(row=idx2, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr(PD, Reset exclusive)
                    cnt2 += 1
                    evb.i2c0_reg16_write(SlaveAddr_0, Addr, write_data0) # 0x55 write
                elif (Addr in exc_list) == True:
                    cnt2 += 1
                    continue

        ##### Ascending READ/WRITE_0 #####
        print("-----ASCENDING READ/WRITE_0 START-----")
        cnt3 = 0
        row2 = 12
        write_data1 = 0xaa
        for idx2 in range(row, len(mode)+1):

            data = ws.cell(row=idx2, column=2).value
            init_val = ws.cell(row=idx2, column=3).value
            reg_name = ws.cell(row=idx2, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr
                    read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr)) # Read 0x55
                    ws2.cell(row=row2, column=col2+4, value=read_data)

                    if int(read_data, 16) == write_data0:
                        ws2.cell(row=row2, column=col2+5, value="PASS")
                        # ws2.cell(row=row2, column=6).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                    else:
                        ws2.cell(row=row2, column=col2+5, value="FAIL")
                        # ws2.cell(row=row2, column=6).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                        cnt3 += 1

                    evb.i2c0_reg16_write(SlaveAddr_0, Addr, write_data1) # 0xaa write

                    row2 += 1

                elif (Addr in exc_list) == True:
                    ws2.cell(row=row2, column=col2+4, value="N/A")
                    ws2.cell(row=row2, column=col2+5, value="N/A")
                    row2 += 1            
            elif data[:2] == "RO":
                continue
            elif data[:2] == "WO":
                continue
            elif data == None:
                break
        ws2.cell(row=len(mode_des)+2, column=col2+5, value=cnt3)

        ##### Ascending READ/WRITE_1 #####
        print("-----ASCENDING READ/WRITE_1 START-----")
        cnt4 = 0
        row2 = 12
        write_data2 = 0x55
        for idx3 in range(row, len(mode)+1):

            data = ws.cell(row=idx3, column=2).value
            init_val = ws.cell(row=idx3, column=3).value
            reg_name = ws.cell(row=idx3, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr
                    read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr)) # Read 0xaa
                    ws2.cell(row=row2, column=col2+6, value=read_data)

                    if int(read_data, 16) == write_data1:
                        ws2.cell(row=row2, column=col2+7, value="PASS")
                        # ws2.cell(row=row2, column=8).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                    else:
                        ws2.cell(row=row2, column=col2+7, value="FAIL")
                        # ws2.cell(row=row2, column=8).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                        cnt4 += 1

                    evb.i2c0_reg16_write(SlaveAddr_0, Addr, write_data2) # ascending write 0x55
                    row2 += 1

                elif (Addr in exc_list) == True:
                    ws2.cell(row=row2, column=col2+6, value="N/A")
                    ws2.cell(row=row2, column=col2+7, value="N/A")
                    row2 += 1            
            elif data[:2] == "RO":
                continue
            elif data[:2] == "WO":
                continue
            elif data == None:
                break
        ws2.cell(row=len(mode_des)+2, column=col2+7, value=cnt4)

        ##### DESCENDING MODE START #####    
        print("-----DESCENDING WRITE/READ_0 START-----")
        cnt6 = 0
        row6 = len(mode_des) # descending start row
        write_data4 = 0xaa
        for idx5 in range(len(mode), row+1, -1):

            data = ws.cell(row=idx5, column=2).value
            init_val = ws.cell(row=idx5, column=3).value
            reg_name = ws.cell(row=idx5, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr

                    read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr)) # Read 0x
                    ws2.cell(row=row6, column=col2+8, value=read_data)
                    if int(read_data, 16) == write_data2:
                        ws2.cell(row=row6, column=col2+9, value="PASS") # column = 12
                        # ws2.cell(row=row6, column=10).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                    else:
                        ws2.cell(row=row6, column=col2+9, value="FAIL")
                        # ws2.cell(row=row6, column=10).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                        cnt6 += 1

                    evb.i2c0_reg16_write(SlaveAddr_0, Addr, write_data4) # descending write 0xaa

                    row6 -= 1

                elif (Addr in exc_list) == True:                
                    ws2.cell(row=row6, column=col2+8, value="N/A")
                    ws2.cell(row=row6, column=col2+9, value="N/A")
                    row6 -= 1
            elif data[:2] == "RO":
                continue
            elif data[:2] == "WO":                
                continue
            elif data == None:            
                break
        ws2.cell(row=len(mode_des)+2, column=col2+9, value=cnt6)
        
        print("-----DESCENDING WRITE/READ_1 START-----")
        cnt7 = 0
        row6 = len(mode_des)
        write_data6 = 0x55
        for idx6 in range(len(mode), row+1, -1):

            data = ws.cell(row=idx6, column=2).value
            init_val = ws.cell(row=idx6, column=3).value
            reg_name = ws.cell(row=idx6, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr                    
                    read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr))            
                    ws2.cell(row=row6, column=col2+10, value=read_data)

                    if int(read_data, 16) == write_data4:
                        ws2.cell(row=row6, column=col2+11, value="PASS")
                        # ws2.cell(row=row6, column=12).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                    else:
                        ws2.cell(row=row6, column=col2+11, value="FAIL")
                        # ws2.cell(row=row6, column=12).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                        cnt7 += 1

                    evb.i2c0_reg16_write(SlaveAddr_0, Addr, write_data6) # descending write 0x55

                    row6 -= 1

                elif (Addr in exc_list) == True:
                    ws2.cell(row=row6, column=col2+10, value="N/A")
                    ws2.cell(row=row6, column=col2+11, value="N/A")
                    row6 -= 1
            elif data[:2] == "RO":                
                continue
            elif data[:2] == "WO":                
                continue
            elif data == None:            
                break
        ws2.cell(row=len(mode_des)+2, column=col2+11, value=cnt7)
        
        print("-----DESCENDING READ_0x55 START-----")
        cnt8 = 0
        row6 = len(mode_des)
        for idx7 in range(len(mode), row+1, -1):

            data = ws.cell(row=idx7, column=2).value
            init_val = ws.cell(row=idx7, column=3).value
            reg_name = ws.cell(row=idx7, column=4).value

            if (data[:2] == "RW") and (init_val != "UNUSED"):
                Addr = int(data[2:], 16)
                if (Addr in exc_list) == False: # int type addr                    
                    read_data = hex(evb.i2c0_reg16_read(SlaveAddr_0, Addr))                    
                    ws2.cell(row=row6, column=col2+12, value=read_data)

                    if int(read_data, 16) == write_data6:
                        ws2.cell(row=row6, column=col2+13, value="PASS")
                        # ws2.cell(row=row6, column=14).fill = PatternFill(fill_type='solid',fgColor=Color('C6EFCE'))
                    else:
                        ws2.cell(row=row6, column=col2+13, value="FAIL")
                        # ws2.cell(row=row6, column=14).fill = PatternFill(fill_type='solid',fgColor=Color('FFC7CD'))
                        cnt8 += 1

                    row6 -= 1

                elif (Addr in exc_list) == True:
                    # cnt8 += 1
                    ws2.cell(row=row6, column=col2+12, value="N/A")
                    ws2.cell(row=row6, column=col2+13, value="N/A")
                    row6 -= 1
            elif data[:2] == "RO":                
                continue
            elif data[:2] == "WO":                
                continue
            elif data == None:            
                break
        ws2.cell(row=len(mode_des)+2, column=col2+13, value=cnt8)
        
        wb2.save(save_path)
        wb2.close()
    print("-----I2C Function Test END-----")

def measure_atest(at, ws, col_atest):

    if "AT0" in at:
        at_addr = 0x002F # AT0 measure
        print("START AT0 ITEMS MEAUSURE")
    elif "AT1" in at:
        at_addr = 0x002E # AT1 measure
        print("START AT1 ITEMS MEAUSURE")

    for idx, items in enumerate(at.items()):
        if idx == 0:
            continue

        test_item = items[0].split("_")[-1]
        # print(idx, "ITEM:",items[0], "KEY:",items[-1][0], "LSL:",items[-1][-1][0], "USL:",items[-1][-1][-1]) # 0, key, value
        ws.cell(row=8, column=col_atest+idx, value=items[0])
        ws.cell(row=row_usl, column=col_atest+idx, value=items[-1][-1][-1])
        ws.cell(row=row_lsl, column=col_atest+idx, value=items[-1][-1][0])
        
        evb.i2c0_reg16_write(SlaveAddr_0, at_addr, items[-1][0]) # ATEST_CON
        time.sleep(1/1000)
    
        if test_item == "VOLTAGE":    
            ws.cell(row=9, column=col_atest+idx, value="V")
            smu.gpib.write(":SOUR:FUNC CURR")
            time.sleep(1/10)
            voltage_val = smu.measure_voltage()
            voltage_read, _, _, _, _ = voltage_val.split(",")
            voltage_read = float(voltage_read)
            ws.cell(row=row_start, column=col_atest+idx, value=voltage_read)
            print(f"{items[0]} : {voltage_read}V")
        elif test_item == "CURRENT":
            ws.cell(row=9, column=col_atest+idx, value="A")
            smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
            time.sleep(1/10)
            current_val = smu.measure_current()
            _, current_read, _, _, _ = current_val.split(",")
            current_read = abs(float(current_read))
            ws.cell(row=row_start, column=col_atest+idx, value=current_read)
            print(f"{items[0]} : {current_read}A")

def measure_atest_all():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START AT0 ALL MEAUSUREMENT-----")
    smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert FRONT Terminal(AT0 Measure)
    time.sleep(1) # Delay 1sec
    print("CONVERT SMU TEMINAL!!!", end="\n\n")

    wb, ws = option_create_sheet()

    at0 = {"AT0":0x002F,\
            "TOP_HVBGR_VOLTAGE":(0x06, (HVBGR_TOP_V_LSL, HVBGR_TOP_V_USL)),\
            "TOP_HVBGR_CURRENT":(0x07, (HVBGR_TOP_I_LSL, HVBGR_TOP_I_USL)),\
            "LDO33_VOLTAGE":(0x08, (LDO33_V_LSL, LDO33_V_USL)),\
            "ADC_REFH_TOP_VOLTAGE":(0x18, (ADC_REFH_TOP_V_LSL, ADC_REFH_TOP_V_USL)),\
            "ADC_REFL_TOP_VOLTAGE":(0x20, (ADC_REFL_TOP_V_LSL, ADC_REFL_TOP_V_USL)),\
            }
    
    measure_atest(at=at0, ws=ws, col_atest=col_start_all_measure)

    print("-----END AT0 MEAUSUREMENT-----")
    smu.gpib.write(":ROUTe:TERMinals REAR") # Convert REAR Terminal(AT1 Measure)
    time.sleep(1) # Delay 1sec
    print("CONVERT SMU TEMINAL!!!", end="\n\n")

    at1 = {"AT1":0x002E,\
            "BTM_HVBGR_VOLTAGE":(0x08, (HVBGR_BTM_V_LSL, HVBGR_BTM_V_USL)),\
            "BTM_HVBGR_CURRENT":(0x10, (HVBGR_BTM_I_LSL, HVBGR_BTM_I_USL)),\
            "ADC_REFH_BTM_VOLTAGE":(0x03, (ADC_REFH_BTM_V_LSL, ADC_REFH_BTM_V_USL)),\
            "ADC_REFL_BTM_VOLTAGE":(0x04, (ADC_REFL_BTM_V_LSL, ADC_REFL_BTM_V_USL)),\
            "MIPIRX_VILCD_VOLTAGE":(0x18, (MIPIRX_VILCD_LSL, MIPIRX_VILCD_USL)),\
            "MIPIRX_VIHCD_VOLTAGE":(0x20, (MIPIRX_VIHCD_LSL, MIPIRX_VIHCD_USL)),\
            "MIPIRX_VULPS_VOLTAGE":(0x28, (MIPIRX_VULPS_LSL, MIPIRX_VULPS_USL)),\
            "MIPIRX_VIL_VOLTAGE":(0x30, (MIPIRX_VIL_LSL, MIPIRX_VIL_USL)),\
            "MIPIRX_VIH_VOLTAGE":(0x38, (MIPIRX_VIH_LSL, MIPIRX_VIH_USL)),\
            }
    
    measure_atest(at=at1, ws=ws, col_atest=col_start_all_measure + 5)
    smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert FRONT Terminal(AT0 Measure)    
    time.sleep(1) # Delay 1sec

    print("CONVERT SMU TEMINAL!!!", end="\n\n")
    print("-----END AT1 MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_hvbgr_top_voltage():
    print("-----START HVBGR TOP VOLTAGE MEAUSUREMENT-----")
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x08) # BTM HVBGR voltage set(AT1)
    # time.sleep(1/2000) # 0.5ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x06) # TOP HVBGR voltage set(AT0)
    smu.gpib.write(":SOUR:FUNC CURR") # SMU current source setting
    time.sleep(1/10) # 0.1s
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"HVBGR_TOP_Voltage:{voltage_read}V")
    print("-----END HVBGR TOP VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_hvbgr_top_current():
    print("-----START HVBGR TOP CURRENT MEAUSUREMENT-----")
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x10) # BTM HVBGR current set(AT1)
    # time.sleep(1/2000) # 0.5ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x07) # TOP HVBGR current set(AT0)
    smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
    smu.set_voltage(0.0) # SMU voltage source volt setting
    time.sleep(1/10) # 0.1s
    current_val = smu.measure_current()
    _, current_read, _, _, _ = current_val.split(",")
    print(f"HVBGR TOP Current:{current_read}A")
    print("-----END HVBGR TOP CURRENT MEAUSUREMENT-----")

    return abs(float(current_read))

def measure_hvbgr_btm_voltage():
    print("-----START HVBGR BTM VOLTAGE MEAUSUREMENT-----")
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x08) # BTM HVBGR voltage set(AT1)
    time.sleep(1/1000) # 0.1ms
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x06) # TOP HVBGR voltage set(AT0)
    smu.gpib.write(":SOUR:FUNC CURR") # SMU current source setting
    time.sleep(1/10) # 0.1s
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"HVBGR_TOP_Voltage:{voltage_read}V")
    print("-----END HVBGR BTM VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_hvbgr_btm_current():
    print("-----START HVBGR BTM CURRENT MEAUSUREMENT-----")
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x10) # BTM HVBGR current set(AT1)
    time.sleep(1/1000) # 0.1ms
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x07) # TOP HVBGR current set(AT0)
    smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
    smu.set_voltage(0.0) # SMU voltage source volt setting
    time.sleep(1/10) # 0.1s
    current_val = smu.measure_current()
    _, current_read, _, _, _ = current_val.split(",")
    print(f"HVBGR BTM Current:{current_read}A")
    print("-----END HVBGR BTM CURRENT MEAUSUREMENT-----")

    return abs(float(current_read))

def measure_ldo33_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START LDO33 VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+7+8+1, value="LDO33_VOLT")
    ws.cell(row=9, column=col_start+7+8+1, value="V")
    ws.cell(row=row_usl, column=col_start+7+8+1, value=LDO33_V_USL)
    ws.cell(row=row_lsl, column=col_start+7+8+1, value=LDO33_V_LSL)

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x08) # LDO33 voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    voltage_read = float(voltage_read)

    ws.cell(row=row_start, column=col_start+7+8+1, value=voltage_read) # row = 12, column = 35

    print(f"LDO33_Voltage:{voltage_read}V")
    print("-----END LDO33 VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_adc_refh_top_voltage():
    print("-----START ADC_REFH TOP VOLTAGE MEAUSUREMENT-----")
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x03) # BTM ADC_REFH voltage(AT1)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x18) # TOP ADC_REFH voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"ADC_TOP_REFH_Voltage:{voltage_read}V")
    print("-----END ADC_REFH TOP VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_adc_refh_btm_voltage():
    print("-----START ADC_REFH BTM VOLTAGE MEAUSUREMENT-----")
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x03) # BTM ADC_REFH voltage(AT1)
    time.sleep(1/1000)
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x18) # TOP ADC_REFH voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"ADC_BTM_REFH_Voltage:{voltage_read}V")
    print("-----END ADC_REFH BTM VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_adc_refl_top_voltage():
    print("-----START ADC_REFL TOP VOLTAGE MEAUSUREMENT-----")
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x04) # BTM ADC_REFL voltage(AT1)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x20) # TOP ADC_REFL voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"ADC_TOP_REFL_Voltage:{voltage_read}V")
    print("-----END ADC_REFL TOP VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_adc_refl_btm_voltage():
    print("-----START ADC_REFL BTM VOLTAGE MEAUSUREMENT-----")
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x04) # BTM ADC_REFL voltage(AT1)
    time.sleep(1/1000)
    # evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x20) # TOP ADC_REFL voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)
    voltage_val = smu.measure_voltage()
    voltage_read, _, _, _, _ = voltage_val.split(",")
    print(f"ADC_REFL_Voltage:{voltage_read}V")
    print("-----END ADC_REFL BTM VOLTAGE MEAUSUREMENT-----")

    return float(voltage_read)

def measure_tempsensor_left_top_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR LEFT TOP VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+41, value="TS_LT_VOLT")
    ws.cell(row=9, column=col_start+41, value="V")
    ws.cell(row=row_usl, column=col_start+41, value="N/A")
    ws.cell(row=row_lsl, column=col_start+41, value="N/A")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x30) # TS LT voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)

    voltage_lst = []
    cnt = 0
    while(cnt < 8):
        cnt += 1
        voltage_val = smu.measure_voltage()
        voltage_read, _, _, _, _ = voltage_val.split(",")
        voltage_read = round(float(voltage_read), 6)
        voltage_lst.append(voltage_read)
        print(f"TS_LT_Voltage:{voltage_read}V")    
    voltage_avg = sum(voltage_lst) / len(voltage_lst)
    voltage_avg = round(float(voltage_avg), 6)
    ws.cell(row=row_start, column=col_start+41, value=voltage_avg)
    
    print(f"TS_LT_Voltage:{voltage_avg}V")
    print("-----END TEMPSENSOR LEFT TOP VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_right_top_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR RIGHT TOP VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+42, value="TS_RT_VOLT")
    ws.cell(row=9, column=col_start+42, value="V")
    ws.cell(row=row_usl, column=col_start+42, value="N/A")
    ws.cell(row=row_lsl, column=col_start+42, value="N/A")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x38) # TS RT voltage(AT0)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)

    voltage_lst = []
    cnt = 0
    while(cnt < 8):
        cnt += 1
        voltage_val = smu.measure_voltage()
        voltage_read, _, _, _, _ = voltage_val.split(",")
        voltage_read = round(float(voltage_read), 6)
        voltage_lst.append(voltage_read)
        print(f"TS_RT_Voltage:{voltage_read}V")    
    voltage_avg = sum(voltage_lst) / len(voltage_lst)
    voltage_avg = round(float(voltage_avg), 6)
    ws.cell(row=row_start, column=col_start+42, value=voltage_avg)

    print(f"TS_RT_Voltage:{voltage_avg}V")
    print("-----END TEMPSENSOR RIGHT TOP VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_left_btm_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR LEFT BTM VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+41, value="TS_LB_VOLT")
    ws.cell(row=9, column=col_start_at1+41, value="V")
    ws.cell(row=row_usl, column=col_start_at1+41, value="N/A")
    ws.cell(row=row_lsl, column=col_start_at1+41, value="N/A")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x06) # TS LB voltage(AT1)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)

    voltage_lst = []
    cnt = 0
    while(cnt < 8):
        cnt += 1
        voltage_val = smu.measure_voltage()
        voltage_read, _, _, _, _ = voltage_val.split(",")
        voltage_read = round(float(voltage_read), 6)
        voltage_lst.append(voltage_read)
        print(f"TS_LB_Voltage:{voltage_read}V")    
    voltage_avg = sum(voltage_lst) / len(voltage_lst)
    voltage_avg = round(float(voltage_avg), 6)
    ws.cell(row=row_start, column=col_start_at1+41, value=voltage_avg)

    print(f"TS_LB_Voltage:{voltage_avg}V")
    print("-----END TEMPSENSOR LEFT BTM VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_right_btm_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR RIGHT BTM VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+42, value="TS_RB_VOLT")
    ws.cell(row=9, column=col_start_at1+42, value="V")
    ws.cell(row=row_usl, column=col_start_at1+42, value="N/A")
    ws.cell(row=row_lsl, column=col_start_at1+42, value="N/A")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, 0x07) # TS RB voltage(AT1)
    smu.gpib.write(":SOUR:FUNC CURR")
    time.sleep(1/10)

    voltage_lst = []
    cnt = 0
    while(cnt < 8):
        cnt += 1
        voltage_val = smu.measure_voltage()
        voltage_read, _, _, _, _ = voltage_val.split(",")
        voltage_read = round(float(voltage_read), 6)
        voltage_lst.append(voltage_read)
        print(f"TS_RB_Voltage:{voltage_read}V")    
    voltage_avg = sum(voltage_lst) / len(voltage_lst)
    voltage_avg = round(float(voltage_avg), 6)
    ws.cell(row=row_start, column=col_start_at1+42, value=voltage_avg)

    print(f"TS_RB_Voltage:{voltage_avg}V")
    print("-----END TEMPSENSOR RIGHT BTM VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_left_top_code():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR LEFT TOP CODE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x01) # ADC_TOP_SOC enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BC, 0x01) # ADC_TOP_ISEL set left top tempsensor
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF0) # TEMP_FILT_MODE:avg
    time.sleep(1/100) # For accurate temp code 10ms
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")
    
    temp_cnt = 11
    temp_af_list = []
    temp_mf_list = []
    for i in range(1, temp_cnt+1, 1): # 1 ~ 11
        if i == 1:
            ws.cell(row=9, column=col_start+42+i, value="CODE(DEC)")
        if i == temp_cnt:
            ws.cell(row=8, column=col_start+42+i, value=f"LT_TEMP_READ_AF(avg)")
            ws.cell(row=12, column=col_start+42+i, value=f"{int(sum(temp_af_list)/len(temp_af_list))}") 
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C5) # TEMPERATURE_LT<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C6) # TEMPERATURE_LT<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        temp_10bit = int(temp_10bit)
        time.sleep(1/1000)
        temp_af_list.append(temp_10bit)

        ws.cell(row=8, column=col_start+42+i, value=f"LT_TEMP_READ{i}_AF")
        ws.cell(row=12, column=col_start+42+i, value=f"{temp_10bit}")       
        ws.cell(row=row_usl, column=col_start+42+i, value=TEMP_CODE_LT_LSL)
        ws.cell(row=row_lsl, column=col_start+42+i, value=TEMP_CODE_LT_USL)
        print(f"LT_TEMPCODE_AF{i} : {temp_10bit}")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF2) # TEMP_FILT_MODE:median
    time.sleep(1/100)
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")

    for i in range(temp_cnt+1, (2*temp_cnt)+1, 1): # 12 ~ 22
        if i == temp_cnt+1:
            ws.cell(row=9, column=col_start+42+i, value="CODE(DEC)")
        if i == (2*temp_cnt):
            ws.cell(row=8, column=col_start+42+i, value=f"LT_TEMP_READ_MF(avg)")
            ws.cell(row=12, column=col_start+42+i, value=f"{int(sum(temp_mf_list)/len(temp_mf_list))}")
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C5) # TEMPERATURE_LT<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C6) # TEMPERATURE_LT<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        time.sleep(1/1000)
        temp_mf_list.append(temp_10bit)

        ws.cell(row=8, column=col_start+42+i, value=f"LT_TEMP_READ{i-11}_MF")
        ws.cell(row=12, column=col_start+42+i, value=f"{temp_10bit}")
        ws.cell(row=row_usl, column=col_start+42+i, value=TEMP_CODE_LT_LSL)
        ws.cell(row=row_lsl, column=col_start+42+i, value=TEMP_CODE_LT_USL)
        print(f"LT_TEMPCODE_MF{i-11} : {temp_10bit}")

    print("-----END TEMPSENSOR LEFT TOP CODE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_right_top_code():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR RIGHT TOP CODE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x01) # ADC_TOP_SOC enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BC, 0x00) # ADC_TOP_ISEL set right top tempsensor
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF0) # TEMP_FILT_MODE:avg
    time.sleep(1/100) # For accurate temp code 10ms
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")
    
    temp_cnt = 11
    temp_af_list = []
    temp_mf_list = []
    for i in range(1, temp_cnt+1, 1): # 1 ~ 11
        if i == 1:
            ws.cell(row=9, column=col_start+42+(2*temp_cnt)+i, value="CODE(DEC)") # 59 -> 81
        if i == temp_cnt:
            ws.cell(row=8, column=col_start+42+(2*temp_cnt)+i, value=f"RT_TEMP_READ_AF(avg)")
            ws.cell(row=12, column=col_start+42+(2*temp_cnt)+i, value=f"{int(sum(temp_af_list)/len(temp_af_list))}") 
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C7) # TEMPERATURE_RT<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C8) # TEMPERATURE_RT<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        temp_10bit = int(temp_10bit)
        time.sleep(1/1000)
        temp_af_list.append(temp_10bit)

        ws.cell(row=8, column=col_start+42+(2*temp_cnt)+i, value=f"RT_TEMP_READ{i}_AF")
        ws.cell(row=12, column=col_start+42+(2*temp_cnt)+i, value=f"{temp_10bit}")       
        ws.cell(row=row_usl, column=col_start+42+(2*temp_cnt)+i, value=TEMP_CODE_RT_LSL)
        ws.cell(row=row_lsl, column=col_start+42+(2*temp_cnt)+i, value=TEMP_CODE_RT_USL)
        print(f"RT_TEMPCODE_AF{i} : {temp_10bit}")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF2) # TEMP_FILT_MODE:median
    time.sleep(1/100)
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")

    for i in range(temp_cnt+1, (2*temp_cnt)+1, 1): # 12 ~ 22
        if i == temp_cnt+1:
            ws.cell(row=9, column=col_start+42+(2*temp_cnt)+i, value="CODE(DEC)")
        if i == (2*temp_cnt):
            ws.cell(row=8, column=col_start+42+(2*temp_cnt)+i, value=f"RT_TEMP_READ_MF(avg)")
            ws.cell(row=12, column=col_start+42+(2*temp_cnt)+i, value=f"{int(sum(temp_mf_list)/len(temp_mf_list))}")
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C7) # TEMPERATURE_LT<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C8) # TEMPERATURE_LT<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        time.sleep(1/1000)
        temp_mf_list.append(temp_10bit)

        ws.cell(row=8, column=col_start+42+(2*temp_cnt)+i, value=f"RT_TEMP_READ{i-11}_MF")
        ws.cell(row=12, column=col_start+42+(2*temp_cnt)+i, value=f"{temp_10bit}")
        ws.cell(row=row_usl, column=col_start+42+(2*temp_cnt)+i, value=TEMP_CODE_RT_LSL)
        ws.cell(row=row_lsl, column=col_start+42+(2*temp_cnt)+i, value=TEMP_CODE_RT_USL)
        print(f"RT_TEMPCODE_MF{i-11} : {temp_10bit}")

    print("-----END TEMPSENSOR RIGHT TOP CODE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_left_btm_code():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR LEFT BTM CODE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x02) # ADC_BTM_SOC enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BC, 0x04) # ADC_BTM_ISEL set left btm tempsensor
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF0) # TEMP_FILT_MODE:avg
    time.sleep(1/100) # For accurate temp code 10ms
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")
    
    temp_cnt = 11
    temp_af_list = []
    temp_mf_list = []
    for i in range(1, temp_cnt+1, 1): # 1 ~ 11
        if i == 1:
            ws.cell(row=9, column=col_start_at1+42+i, value="CODE(DEC)")
        if i == temp_cnt:
            ws.cell(row=8, column=col_start_at1+42+i, value=f"LB_TEMP_READ_AF(avg)")
            ws.cell(row=12, column=col_start_at1+42+i, value=f"{int(sum(temp_af_list)/len(temp_af_list))}") 
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C9) # TEMPERATURE_LB<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CA) # TEMPERATURE_LB<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        temp_10bit = int(temp_10bit)
        time.sleep(1/1000)
        temp_af_list.append(temp_10bit)

        ws.cell(row=8, column=col_start_at1+42+i, value=f"LB_TEMP_READ{i}_AF")
        ws.cell(row=12, column=col_start_at1+42+i, value=f"{temp_10bit}")       
        ws.cell(row=row_usl, column=col_start_at1+42+i, value=TEMP_CODE_LB_LSL)
        ws.cell(row=row_lsl, column=col_start_at1+42+i, value=TEMP_CODE_LB_USL)
        print(f"LB_TEMPCODE_AF{i} : {temp_10bit}")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF2) # TEMP_FILT_MODE:median
    time.sleep(1/100)
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")

    for i in range(temp_cnt+1, (2*temp_cnt)+1, 1): # 12 ~ 22
        if i == temp_cnt+1:
            ws.cell(row=9, column=col_start_at1+42+i, value="CODE(DEC)")
        if i == (2*temp_cnt):
            ws.cell(row=8, column=col_start_at1+42+i, value=f"LB_TEMP_READ_MF(avg)")
            ws.cell(row=12, column=col_start_at1+42+i, value=f"{int(sum(temp_mf_list)/len(temp_mf_list))}")
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C9) # TEMPERATURE_LB<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CA) # TEMPERATURE_LB<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        time.sleep(1/1000)
        temp_mf_list.append(temp_10bit)

        ws.cell(row=8, column=col_start_at1+42+i, value=f"LB_TEMP_READ{i-11}_MF")
        ws.cell(row=12, column=col_start_at1+42+i, value=f"{temp_10bit}")
        ws.cell(row=row_usl, column=col_start_at1+42+i, value=TEMP_CODE_LB_LSL)
        ws.cell(row=row_lsl, column=col_start_at1+42+i, value=TEMP_CODE_LB_USL)
        print(f"LB_TEMPCODE_MF{i-11} : {temp_10bit}")

    print("-----END TEMPSENSOR LEFT BTM CODE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_tempsensor_right_btm_code():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START TEMPSENSOR RIGHT BTM CODE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x02) # ADC_BTM_SOC enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BC, 0x00) # ADC_BTM_ISEL set right btm tempsensor
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF0) # TEMP_FILT_MODE:avg
    time.sleep(1/100) # For accurate temp code 10ms
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")
    
    temp_cnt = 11
    temp_af_list = []
    temp_mf_list = []
    for i in range(1, temp_cnt+1, 1): # 1 ~ 11
        if i == 1:
            ws.cell(row=9, column=col_start_at1+42+(2*temp_cnt)+i, value="CODE(DEC)") # 59 -> 81
        if i == temp_cnt:
            ws.cell(row=8, column=col_start_at1+42+(2*temp_cnt)+i, value=f"RB_TEMP_READ_AF(avg)")
            ws.cell(row=12, column=col_start_at1+42+(2*temp_cnt)+i, value=f"{int(sum(temp_af_list)/len(temp_af_list))}") 
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CB) # TEMPERATURE_RB<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CC) # TEMPERATURE_RB<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        temp_10bit = int(temp_10bit)
        time.sleep(1/1000)
        temp_af_list.append(temp_10bit)

        ws.cell(row=8, column=col_start_at1+42+(2*temp_cnt)+i, value=f"RB_TEMP_READ{i}_AF")
        ws.cell(row=12, column=col_start_at1+42+(2*temp_cnt)+i, value=f"{temp_10bit}")       
        ws.cell(row=row_usl, column=col_start_at1+42+(2*temp_cnt)+i, value=TEMP_CODE_RB_LSL)
        ws.cell(row=row_lsl, column=col_start_at1+42+(2*temp_cnt)+i, value=TEMP_CODE_RB_USL)
        print(f"RB_TEMPCODE_AF{i} : {temp_10bit}")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF2) # TEMP_FILT_MODE:median
    time.sleep(1/100)
    print(f"FILTMODE : {hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05D1))}")

    for i in range(temp_cnt+1, (2*temp_cnt)+1, 1): # 12 ~ 22
        if i == temp_cnt+1:
            ws.cell(row=9, column=col_start_at1+42+(2*temp_cnt)+i, value="CODE(DEC)")
        if i == (2*temp_cnt):
            ws.cell(row=8, column=col_start_at1+42+(2*temp_cnt)+i, value=f"RB_TEMP_READ_MF(avg)")
            ws.cell(row=12, column=col_start_at1+42+(2*temp_cnt)+i, value=f"{int(sum(temp_mf_list)/len(temp_mf_list))}")
            break
        
        temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CB) # TEMPERATURE_RB<9:2>
        time.sleep(1/1000)
        temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CC) # TEMPERATURE_RB<1:0>
        temp_10bit = (temp_msb8<<2) + temp_lsb2
        time.sleep(1/1000)
        temp_mf_list.append(temp_10bit)

        ws.cell(row=8, column=col_start_at1+42+(2*temp_cnt)+i, value=f"RB_TEMP_READ{i-11}_MF")
        ws.cell(row=12, column=col_start_at1+42+(2*temp_cnt)+i, value=f"{temp_10bit}")
        ws.cell(row=row_usl, column=col_start_at1+42+(2*temp_cnt)+i, value=TEMP_CODE_RB_LSL)
        ws.cell(row=row_lsl, column=col_start_at1+42+(2*temp_cnt)+i, value=TEMP_CODE_RB_USL)
        print(f"RB_TEMPCODE_MF{i-11} : {temp_10bit}")

    print("-----END TEMPSENSOR RIGHT BTM CODE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_sub_pixel_current():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START SUB-PIXEL CURRENT MEAUSUREMENT-----")
    wb, ws = option_create_sheet()
    col_sub_pixel = col_start + 87
    ws.cell(row=8, column=col_sub_pixel, value="SUB-PIXEL_DATA")
    ws.cell(row=9, column=col_sub_pixel, value="CODE(DEC)")
    ws.cell(row=8, column=col_sub_pixel+1, value="SUB-PIXEL_CURR") 
    ws.cell(row=9, column=col_sub_pixel+1, value="A")
    ws.cell(row=8, column=col_sub_pixel+1+1, value="MAX_SUB-PIXEL_CURR")
    ws.cell(row=9, column=col_sub_pixel+1+1, value="A")
    ws.cell(row=row_usl, column=col_sub_pixel+1+1, value=MAX_SUB_PIXEL_I_USL)
    ws.cell(row=row_lsl, column=col_sub_pixel+1+1, value=MAX_SUB_PIXEL_I_LSL)

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002f, 0x45) # Sub-pixel current setting(AT0)
    smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
    smu.set_voltage(0.1) # Set AT0:0.1V
    smu.gpib.write(":SENS:CURR:PROT 300E-6") # Set Cmpl 300uA
    time.sleep(1/10) # 0.1s

    # col_data_sweep
    N_start = 0
    N_finish = 4095 # data bit
    N_step = 1
    col_test_addr = [187]
    row_sub_pixel = 12
    N_sub_pixel = 300

    for col_test in col_test_addr:
        ws.cell(row=row_sub_pixel, column=col_sub_pixel, value=col_test)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f0, 0x01) # col_test_en
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f1, col_test>>8) # col_test_addr<8>
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f2, col_test) # col_test_addr<7:0>
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x00) # col_test_trig
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x01) # col_test_trig
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x00) # col_test_trig

        for data in range(N_start, N_finish+1, N_step):
            cnt = 0
            evb.i2c0_reg16_write(SlaveAddr_0, 0x00d6, 0x80) # col_man_en
            evb.i2c0_reg16_write(SlaveAddr_0, 0x00d6, (data>>8) + 128) # col_man_en & col_data_man<11:8>
            time.sleep(1/1000) # 1ms
            evb.i2c0_reg16_write(SlaveAddr_0, 0x00d7, data) # col_data_man<7:0>
            
            if data <= 100 or data == N_finish:
                data_current = []
                while(cnt < 16):
                    cnt += 1
                    current_val = smu.measure_current()
                    _, current_read, _, _, _ = current_val.split(",")
                    current_read = abs(float(current_read))
                    data_current.append(current_read)
                    print("CNT:",cnt,"DATA:", data, current_read)
                    if cnt == 16:
                        current_avg = sum(data_current) / len(data_current)
                        if data == N_finish:
                            ws.cell(row=12, column=col_sub_pixel+1+1, value=current_avg / N_sub_pixel) # Maximum_sub-pixel current
                            
                            sweep_lst = [i for i in range(0, 8, 1)]
                            row_icon, row_mcon = 12
                            for icon in sweep_lst:
                                evb.i2c0_reg16_write(SlaveAddr_0, 0x0418, icon) # icon sweep
                                time.sleep(1/100)
                                current_val = smu.measure_current()
                                _, current_read, _, _, _ = current_val.split(",")
                                current_read = abs(float(current_read))
                                ws.cell(row=8, column=246, value="GV64_ICON SWEEP")
                                ws.cell(row=9, column=246, value="A")
                                ws.cell(row=row_icon, column=246, value=current_read / N_sub_pixel)
                                row_icon += 1
                            for mcon in list(map(lambda x:(x*16)+1, sweep_lst)):
                                evb.i2c0_reg16_write(SlaveAddr_0, 0x0418, mcon) # mcon sweep
                                time.sleep(1/100)
                                current_val = smu.measure_current()
                                _, current_read, _, _, _ = current_val.split(",")
                                current_read = abs(float(current_read))
                                ws.cell(row=8, column=247, value="GV64_MCON SWEEP")
                                ws.cell(row=9, column=247, value="A")
                                ws.cell(row=row_mcon, column=247, value=current_read / N_sub_pixel)   
                                row_mcon += 1

                        ws.cell(row=row_sub_pixel, column=col_sub_pixel, value=data) # row = 12, col = 103
                        ws.cell(row=row_sub_pixel, column=col_sub_pixel+1, value=current_avg / N_sub_pixel) # col = 104
            else:
                if data == 1500:
                    smu.gpib.write(":SENS:CURR:PROT 10E-3") # Set Cmpl 5mA
                    time.sleep(1/10) # 0.1s
                current_val = smu.measure_current()
                _, current_read, _, _, _ = current_val.split(",")
                current_read = abs(float(current_read))
                print("DATA:", data, current_read)
                ws.cell(row=row_sub_pixel, column=col_sub_pixel, value=data) # row = 12, col = 103
                ws.cell(row=row_sub_pixel, column=col_sub_pixel+1, value=current_read / N_sub_pixel) # col = 104                                     
                # if data == N_finish:
                #     ws.cell(row=12, column=col_sub_pixel+1+1, value=current_read / N_sub_pixel) # Maximum_sub-pixel current
            
            print(f"{col_test}_test_column_data({data})_output_current : {current_read}A")
            row_sub_pixel += 1
        col_sub_pixel += 1 # In case of test_address sweep
    
    print("-----END SUB-PIXEL CURRENT MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def calibration_hvbgr_top_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START HVBGR TOP VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+0, value="HVBGR_TOP_VOLT")
    ws.cell(row=8, column=col_start+1, value="HVBGR_TOP_V_INDCATOR")
    ws.cell(row=8, column=col_start+2, value="BIAS_TOP_VCON")
    ws.cell(row=8, column=col_start+3, value="HVBGR_TOP_V_CAL1")
    ws.cell(row=8, column=col_start+4, value="BIAS_TOP_VCON")
    ws.cell(row=8, column=col_start+5, value="HVBGR_TOP_V_CAL2")
    ws.cell(row=8, column=col_start+6, value="BIAS_TOP_VCON")
    ws.cell(row=8, column=col_start+7, value="HVBGR_TOP_V_CAL3")
    ws.cell(row=9, column=col_start+0, value="V")
    ws.cell(row=row_usl, column=col_start+0, value=HVBGR_TOP_V_USL)
    ws.cell(row=row_lsl, column=col_start+0, value=HVBGR_TOP_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_BIAS_TOP_VCON = 3 # BIAS_TOP_VCON control bit - 1
    top_hvbgr_voltage = measure_hvbgr_top_voltage()
    ws.cell(row=row_start, column=col_start, value=top_hvbgr_voltage) # row = 12, column = 20
    print(f"HVBGR Voltage{cnt}:{top_hvbgr_voltage}V")
    
    if HVBGR_TOP_V_LSL <= top_hvbgr_voltage <= HVBGR_TOP_V_USL:    
        addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012))
        addr0012 = int(addr0012, 16)
        global BIAS_TOP_VCON
        BIAS_TOP_VCON = addr0012>>4 & 0x0f # TOP HVBGR VCON<3:0>, RCON<3:0> upper 4bit
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start+1, value=INDICATOR) # row = 12, column = 21
        while(cnt < 3):
            if top_hvbgr_voltage < HVBGR_TOP_V_LSL:
                cnt += 1
                DELTA_TOP_VCON = 0
                DELTA_TOP_VCON -= (2**(N_BIAS_TOP_VCON-cnt)) # 4, 2, 1
                addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)) # Delta_vcon write
                addr0012 = int(addr0012, 16)
                BIAS_TOP_VCON = addr0012>>4 & 0x0f # upper 4bit
                BIAS_TOP_VCON += DELTA_TOP_VCON
                
                ws.cell(row=row_start, column=(col_start+0) + (2 * cnt), value=BIAS_TOP_VCON) # row = 12, column = 22
                print(f"DOWN, cnt:{cnt}, DELTA_TOP_VCON:{DELTA_TOP_VCON}, BIAS_TOP_VCON:{BIAS_TOP_VCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0012, (BIAS_TOP_VCON<<4) + 8) # TOP HVBGR voltage
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)))
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+0) + (2 * cnt) + 1, value=voltage_read) # row = 12, column = 23
                print(f"HVBGR Voltage{cnt}:{voltage_read}V")
                
                top_hvbgr_voltage = voltage_read
                
                if HVBGR_TOP_V_LSL <= voltage_read <= HVBGR_TOP_V_USL:
                    break

            elif top_hvbgr_voltage > HVBGR_TOP_V_LSL:
                cnt += 1
                DELTA_TOP_VCON = 0
                DELTA_TOP_VCON += (2**(N_BIAS_TOP_VCON-cnt)) # 4, 2, 1
                addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012))
                addr0012 = int(addr0012, 16)
                BIAS_TOP_VCON = addr0012>>4 & 0x0f # upper 4bit
                BIAS_TOP_VCON += DELTA_TOP_VCON

                ws.cell(row=row_start, column=(col_start+0) + (2 * cnt), value=BIAS_TOP_VCON) # row = 12, column = 22
                print(f"UP, cnt:{cnt}, DELTA_TOP_VCON:{DELTA_TOP_VCON}, BIAS_TOP_VCON:{BIAS_TOP_VCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0012, (BIAS_TOP_VCON<<4) + 8) # TOP HVBGR voltage
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)))
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+0) + (2 * cnt) + 1, value=voltage_read) # row = 12, column = 23
                print(f"HVBGR Voltage{cnt}:{voltage_read}V")
                
                top_hvbgr_voltage = voltage_read
                
                if HVBGR_TOP_V_LSL <= voltage_read <= HVBGR_TOP_V_USL:
                    break
    
    print(f"HVBGR_TOP_V_CAL_CNT:{cnt}")
    print("-----END HVBGR TOP VOLTAGE CALIBRATION-----")
    
    wb.save(save_path)
    wb.close()

def calibration_hvbgr_btm_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START HVBGR BTM VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+0, value="HVBGR_BTM_VOLT")
    ws.cell(row=8, column=col_start_at1+1, value="HVBGR_BTM_V_INDCATOR")
    ws.cell(row=8, column=col_start_at1+2, value="BIAS_BTM_VCON")
    ws.cell(row=8, column=col_start_at1+3, value="HVBGR_BTM_V_CAL1")
    ws.cell(row=8, column=col_start_at1+4, value="BIAS_BTM_VCON")
    ws.cell(row=8, column=col_start_at1+5, value="HVBGR_BTM_V_CAL2")
    ws.cell(row=8, column=col_start_at1+6, value="BIAS_BTM_VCON")
    ws.cell(row=8, column=col_start_at1+7, value="HVBGR_BTM_V_CAL3")
    ws.cell(row=9, column=col_start_at1+0, value="V")
    ws.cell(row=row_usl, column=col_start_at1+0, value=HVBGR_BTM_V_USL)
    ws.cell(row=row_lsl, column=col_start_at1+0, value=HVBGR_BTM_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_BIAS_BTM_VCON = 3 # BIAS_BTM_VCON control bit - 1
    top_hvbgr_voltage = measure_hvbgr_btm_voltage()
    ws.cell(row=row_start, column=col_start_at1, value=top_hvbgr_voltage)
    print(f"HVBGR Voltage{cnt}:{top_hvbgr_voltage}V")
    
    if HVBGR_BTM_V_LSL <= top_hvbgr_voltage <= HVBGR_BTM_V_USL:    
        addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013))
        addr0013 = int(addr0013, 16)
        global BIAS_BTM_VCON
        BIAS_BTM_VCON = addr0013>>4 & 0x0f # TOP HVBGR VCON<3:0>, RCON<3:0> upper 4bit
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start_at1+1, value=INDICATOR) 
        while(cnt < 3):
            if top_hvbgr_voltage < HVBGR_BTM_V_LSL:
                cnt += 1
                DELTA_BTM_VCON = 0
                DELTA_BTM_VCON -= (2**(N_BIAS_BTM_VCON-cnt)) # 4, 2, 1
                addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)) # Delta_vcon write
                addr0013 = int(addr0013, 16)
                BIAS_BTM_VCON = addr0013>>4 & 0x0f # upper 4bit
                BIAS_BTM_VCON += DELTA_BTM_VCON
                
                ws.cell(row=row_start, column=(col_start_at1+0) + (2 * cnt), value=BIAS_BTM_VCON)
                print(f"DOWN, cnt:{cnt}, DELTA_BTM_VCON:{DELTA_BTM_VCON}, BIAS_BTM_VCON:{BIAS_BTM_VCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0013, (BIAS_BTM_VCON<<4) + 8) # BTM HVBGR voltage
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)))
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+0) + (2 * cnt) + 1, value=voltage_read)
                print(f"HVBGR Voltage{cnt}:{voltage_read}V")
                
                top_hvbgr_voltage = voltage_read
                
                if HVBGR_BTM_V_LSL <= voltage_read <= HVBGR_BTM_V_USL:
                    break

            elif top_hvbgr_voltage > HVBGR_BTM_V_LSL:
                cnt += 1
                DELTA_BTM_VCON = 0
                DELTA_BTM_VCON += (2**(N_BIAS_BTM_VCON-cnt)) # 4, 2, 1
                addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013))
                addr0013 = int(addr0013, 16)
                BIAS_BTM_VCON = addr0013>>4 & 0x0f # upper 4bit
                BIAS_BTM_VCON += DELTA_BTM_VCON

                ws.cell(row=row_start, column=(col_start_at1+0) + (2 * cnt), value=BIAS_BTM_VCON)
                print(f"UP, cnt:{cnt}, DELTA_BTM_VCON:{DELTA_BTM_VCON}, BIAS_BTM_VCON:{BIAS_BTM_VCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0013, (BIAS_BTM_VCON<<4) + 8) # BTM HVBGR voltage
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)))
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+0) + (2 * cnt) + 1, value=voltage_read)
                print(f"HVBGR Voltage{cnt}:{voltage_read}V")
                
                top_hvbgr_voltage = voltage_read
                
                if HVBGR_BTM_V_LSL <= voltage_read <= HVBGR_BTM_V_USL:
                    break
    
    print(f"HVBGR_BTM_V_CAL_CNT:{cnt}")
    print("-----END HVBGR BTM VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_hvbgr_top_current():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START HVBGR TOP CURRENT CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+0+8, value="HVBGR_TOP_CURR")
    ws.cell(row=8, column=col_start+1+8, value="HVBGR_TOP_I_INDCATOR")
    ws.cell(row=8, column=col_start+2+8, value="BIAS_TOP_RCON")
    ws.cell(row=8, column=col_start+3+8, value="HVBGR_TOP_I_CAL1")
    ws.cell(row=8, column=col_start+4+8, value="BIAS_TOP_RCON")
    ws.cell(row=8, column=col_start+5+8, value="HVBGR_TOP_I_CAL2")
    ws.cell(row=8, column=col_start+6+8, value="BIAS_TOP_RCON")
    ws.cell(row=8, column=col_start+7+8, value="HVBGR_TOP_I_CAL3")
    ws.cell(row=9, column=col_start+0+8, value="A")
    ws.cell(row=row_usl, column=col_start+0+8, value=HVBGR_TOP_I_USL)
    ws.cell(row=row_lsl, column=col_start+0+8, value=HVBGR_TOP_I_LSL)

    cnt = 0
    INDICATOR = 0
    N_BIAS_TOP_RCON = 3 # BIAS_TOP_RCON control bit - 1
    top_hvbgr_current = measure_hvbgr_top_current()
    ws.cell(row=row_start, column=col_start+8, value=top_hvbgr_current) # row = 12, column = 20
    print(f"HVBGR CURRENT{cnt}:{top_hvbgr_current}A")

    if HVBGR_TOP_I_LSL <= top_hvbgr_current <= HVBGR_TOP_I_USL:
        addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012))
        addr0012 = int(addr0012, 16)
        global BIAS_TOP_RCON
        BIAS_TOP_RCON = addr0012 & 0x0f # lower 4bit
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start+1+8, value=INDICATOR)
        while(cnt < 3):
            if top_hvbgr_current < HVBGR_TOP_I_LSL:
                cnt += 1
                DELTA_TOP_RCON = 0
                DELTA_TOP_RCON += (2**(N_BIAS_TOP_RCON-cnt)) # 4, 2, 1
                addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)) # Delta_vcon write
                addr0012 = int(addr0012, 16)
                BIAS_TOP_RCON = addr0012 & 0x0f # lower 4bit
                BIAS_TOP_RCON += DELTA_TOP_RCON

                ws.cell(row=row_start, column=(col_start+0+8) + (2 * cnt), value=BIAS_TOP_RCON)
                print(f"DOWN, cnt:{cnt}, DELTA_TOP_RCON:{DELTA_TOP_RCON}, BIAS_TOP_RCON:{BIAS_TOP_RCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0012, (BIAS_TOP_RCON) + (BIAS_TOP_VCON<<4)) # BTM HVBGR current
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)))
                
                smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
                time.sleep(0.1)
                current_val = smu.measure_current()
                _, current_read, _, _, _ = current_val.split(",")
                current_read = abs(float(current_read))
                ws.cell(row=row_start, column=(col_start+0+8) + (2 * cnt) + 1, value=current_read)
                print(f"HVBGR CURRENT{cnt}:{current_read}A")
                
                top_hvbgr_current = current_read
                
                if HVBGR_TOP_I_LSL <= current_read <= HVBGR_TOP_I_USL:
                    break

            elif top_hvbgr_current > HVBGR_TOP_I_LSL:
                cnt += 1
                DELTA_TOP_RCON = 0
                DELTA_TOP_RCON -= (2**(N_BIAS_TOP_RCON-cnt)) # 4, 2, 1
                addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012))
                addr0012 = int(addr0012, 16)
                BIAS_TOP_RCON = addr0012 & 0x0f # lower 4bit
                BIAS_TOP_RCON += DELTA_TOP_RCON

                ws.cell(row=row_start, column=(col_start+0+8) + (2 * cnt), value=BIAS_TOP_RCON)
                print(f"UP, cnt:{cnt}, DELTA_TOP_RCON:{DELTA_TOP_RCON}, BIAS_TOP_RCON:{BIAS_TOP_RCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0012, (BIAS_TOP_RCON) + (BIAS_TOP_VCON<<4)) # BTM HVBGR current
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)))
                
                smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
                time.sleep(0.1)
                current_val = smu.measure_current()
                _, current_read, _, _, _ = current_val.split(",")
                current_read = abs(float(current_read))
                ws.cell(row=row_start, column=(col_start+0+8) + (2 * cnt) + 1, value=current_read)
                print(f"HVBGR CURRENT{cnt}:{current_read}A")

                top_hvbgr_current = current_read

                if HVBGR_TOP_I_LSL <= current_read <= HVBGR_TOP_I_USL:
                    break
    
    print(f"HVBGR_TOP_I_CAL_CNT:{cnt}")
    print("-----END HVBGR TOP VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_hvbgr_btm_current():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START HVBGR BTM CURRENT CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+0+8, value="HVBGR_BTM_CURR")
    ws.cell(row=8, column=col_start_at1+1+8, value="HVBGR_BTM_I_INDCATOR")
    ws.cell(row=8, column=col_start_at1+2+8, value="BIAS_BTM_RCON")
    ws.cell(row=8, column=col_start_at1+3+8, value="HVBGR_BTM_I_CAL1")
    ws.cell(row=8, column=col_start_at1+4+8, value="BIAS_BTM_RCON")
    ws.cell(row=8, column=col_start_at1+5+8, value="HVBGR_BTM_I_CAL2")
    ws.cell(row=8, column=col_start_at1+6+8, value="BIAS_BTM_RCON")
    ws.cell(row=8, column=col_start_at1+7+8, value="HVBGR_BTM_I_CAL3")
    ws.cell(row=9, column=col_start_at1+0+8, value="A")
    ws.cell(row=row_usl, column=col_start_at1+0+8, value=HVBGR_BTM_I_USL)
    ws.cell(row=row_lsl, column=col_start_at1+0+8, value=HVBGR_BTM_I_LSL)

    cnt = 0
    INDICATOR = 0
    N_BIAS_BTM_RCON = 3 # BIAS_BTM_RCON control bit - 1
    top_hvbgr_current = measure_hvbgr_btm_current()
    ws.cell(row=row_start, column=col_start_at1+8, value=top_hvbgr_current)
    print(f"HVBGR CURRENT{cnt}:{top_hvbgr_current}A")

    if HVBGR_BTM_I_LSL <= top_hvbgr_current <= HVBGR_BTM_I_USL:
        addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013))
        addr0013 = int(addr0013, 16)
        global BIAS_BTM_RCON
        BIAS_BTM_RCON = addr0013 & 0x0f # lower 4bit
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start_at1+1+8, value=INDICATOR)
        while(cnt < 3):
            if top_hvbgr_current < HVBGR_BTM_I_LSL:
                cnt += 1
                DELTA_BTM_RCON = 0
                DELTA_BTM_RCON += (2**(N_BIAS_BTM_RCON-cnt)) # 4, 2, 1
                addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)) # Delta_vcon write
                addr0013 = int(addr0013, 16)
                BIAS_BTM_RCON = addr0013 & 0x0f # lower 4bit
                BIAS_BTM_RCON += DELTA_BTM_RCON

                ws.cell(row=row_start, column=(col_start_at1+0+8) + (2 * cnt), value=BIAS_BTM_RCON)
                print(f"DOWN, cnt:{cnt}, DELTA_BTM_RCON:{DELTA_BTM_RCON}, BIAS_BTM_RCON:{BIAS_BTM_RCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0013, (BIAS_BTM_RCON) + (BIAS_BTM_VCON<<4)) # BTM HVBGR current
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)))
                
                smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
                time.sleep(0.1)
                current_val = smu.measure_current()
                _, current_read, _, _, _ = current_val.split(",")
                current_read = abs(float(current_read))
                ws.cell(row=row_start, column=(col_start_at1+0+8) + (2 * cnt) + 1, value=current_read)
                print(f"HVBGR CURRENT{cnt}:{current_read}A")
                
                top_hvbgr_current = current_read
                
                if HVBGR_BTM_I_LSL <= current_read <= HVBGR_BTM_I_USL:
                    break

            elif top_hvbgr_current > HVBGR_BTM_I_LSL:
                cnt += 1
                DELTA_BTM_RCON = 0
                DELTA_BTM_RCON -= (2**(N_BIAS_BTM_RCON-cnt)) # 4, 2, 1
                addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013))
                addr0013 = int(addr0013, 16)
                BIAS_BTM_RCON = addr0013 & 0x0f # lower 4bit
                BIAS_BTM_RCON += DELTA_BTM_RCON

                ws.cell(row=row_start, column=(col_start_at1+0+8) + (2 * cnt), value=BIAS_BTM_RCON)
                print(f"UP, cnt:{cnt}, DELTA_BTM_RCON:{DELTA_BTM_RCON}, BIAS_BTM_RCON:{BIAS_BTM_RCON}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x0013, (BIAS_BTM_RCON) + (BIAS_BTM_VCON<<4)) # BTM HVBGR current
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)))
                
                smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
                time.sleep(0.1)
                current_val = smu.measure_current()
                _, current_read, _, _, _ = current_val.split(",")
                current_read = abs(float(current_read))
                ws.cell(row=row_start, column=(col_start_at1+0+8) + (2 * cnt) + 1, value=current_read)
                print(f"HVBGR CURRENT{cnt}:{current_read}A")

                top_hvbgr_current = current_read

                if HVBGR_BTM_I_LSL <= current_read <= HVBGR_BTM_I_USL:
                    break
    
    print(f"HVBGR_BTM_I_CAL_CNT:{cnt}")
    print("-----END HVBGR BTM CURRENT CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_adc_refh_top_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START ADC_REFH TOP VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+0+7+8+1+1, value="ADC_REFH_TOP_VOLT")
    ws.cell(row=8, column=col_start+1+7+8+1+1, value="ADC_REFH_TOP_V_INDCATOR")
    ws.cell(row=8, column=col_start+2+7+8+1+1, value="ADC_TOP_VHGEN<4:0>")
    ws.cell(row=8, column=col_start+3+7+8+1+1, value="ADC_REFH_TOP_V_CAL1")
    ws.cell(row=8, column=col_start+4+7+8+1+1, value="ADC_TOP_VHGEN<4:0>")
    ws.cell(row=8, column=col_start+5+7+8+1+1, value="ADC_REFH_TOP_V_CAL2")
    ws.cell(row=8, column=col_start+6+7+8+1+1, value="ADC_TOP_VHGEN<4:0>")
    ws.cell(row=8, column=col_start+7+7+8+1+1, value="ADC_REFH_TOP_V_CAL3")
    ws.cell(row=8, column=col_start+8+7+8+1+1, value="ADC_TOP_VHGEN<4:0>")
    ws.cell(row=8, column=col_start+9+7+8+1+1, value="ADC_REFH_TOP_V_CAL4")
    ws.cell(row=8, column=col_start+10+7+8+1+1, value="ADC_TOP_VHGEN<4:0>") # added 47
    ws.cell(row=8, column=col_start+11+7+8+1+1, value="ADC_REFH_TOP_V_CAL5") # added 48
    ws.cell(row=9, column=col_start+0+7+8+1+1, value="V")
    ws.cell(row=row_usl, column=col_start+0+7+8+1+1, value=ADC_REFH_TOP_V_USL)
    ws.cell(row=row_lsl, column=col_start+0+7+8+1+1, value=ADC_REFH_TOP_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_ADC_REFH_TOP_VHGEN = 4 # control bit - 1
    global ADC_REFH_TOP_VHGEN
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BD, 0x10)
    time.sleep(1/1000)
    top_adc_refh_voltage = measure_adc_refh_top_voltage()
    ws.cell(row=row_start, column=col_start+0+7+8+1+1, value=top_adc_refh_voltage) # 37
    print(f"ADC_REFH_TOP Voltage{cnt}:{top_adc_refh_voltage}V")
    
    if ADC_REFH_TOP_V_LSL <= top_adc_refh_voltage <= ADC_REFH_TOP_V_USL:    
        addr05BD = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD))
        addr05BD = int(addr05BD, 16)
        ADC_REFH_TOP_VHGEN = addr05BD
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start+0+7+8+1+1+1, value=INDICATOR) # 38
        while(cnt < 5):
            if top_adc_refh_voltage < ADC_REFH_TOP_V_LSL:
                if cnt == 0:
                    DELTA_TOP_VHGEN = 15
                elif cnt >= 1:
                    DELTA_TOP_VHGEN = 0
                    DELTA_TOP_VHGEN += (2**(N_ADC_REFH_TOP_VHGEN-cnt)) #8, 4, 2, 1
                
                addr05BD = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD))
                addr05BD = int(addr05BD, 16)
                ADC_REFH_TOP_VHGEN = addr05BD
                ADC_REFH_TOP_VHGEN += DELTA_TOP_VHGEN
                ws.cell(row=row_start, column=(col_start+0+7+8+1+1) + (2 * cnt) + 2, value=ADC_REFH_TOP_VHGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_TOP_VHGEN:{DELTA_TOP_VHGEN}, ADC_REFH_TOP_VHGEN:{ADC_REFH_TOP_VHGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BD, ADC_REFH_TOP_VHGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD)))
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+0+7+8+1+1) + (2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFH_TOP Voltage{cnt}:{voltage_read}V")
                
                cnt += 1    
                
                top_adc_refh_voltage = voltage_read                
                if ADC_REFH_TOP_V_LSL <= voltage_read <= ADC_REFH_TOP_V_USL:
                    break
            elif top_adc_refh_voltage > ADC_REFH_TOP_V_LSL:
                DELTA_TOP_VHGEN = 0
                DELTA_TOP_VHGEN -= (2**(N_ADC_REFH_TOP_VHGEN-cnt)) # 8, 4, 2, 1
                addr05BD = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD))
                addr05BD = int(addr05BD, 16)
                ADC_REFH_TOP_VHGEN = addr05BD
                ADC_REFH_TOP_VHGEN += DELTA_TOP_VHGEN

                ws.cell(row=row_start, column=(col_start+0+7+8+1+1) + (2 * cnt) + 2, value=ADC_REFH_TOP_VHGEN)
                print(f"UP, cnt:{cnt}, DELTA_TOP_VHGEN:{DELTA_TOP_VHGEN}, ADC_REFH_TOP_VHGEN:{ADC_REFH_TOP_VHGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BD, ADC_REFH_TOP_VHGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD)))                            
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+0+7+8+1+1)+(2 * cnt)+ 3, value=voltage_read)
                print(f"ADC_REFH_TOP Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                top_adc_refh_voltage = voltage_read
                if ADC_REFH_TOP_V_LSL <= voltage_read <= ADC_REFH_TOP_V_USL:
                    break
    
    print(f"ADC_REFH_TOP_V_CAL_CNT:{cnt}")
    print("-----END ADC_REFH TOP VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_adc_refh_btm_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START ADC_REFH BTM VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+0+7+8+1, value="ADC_REFH_BTM_VOLT")
    ws.cell(row=8, column=col_start_at1+1+7+8+1, value="ADC_REFH_BTM_V_INDCATOR")
    ws.cell(row=8, column=col_start_at1+2+7+8+1, value="ADC_BTM_VHGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+3+7+8+1, value="ADC_REFH_BTM_V_CAL1")
    ws.cell(row=8, column=col_start_at1+4+7+8+1, value="ADC_BTM_VHGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+5+7+8+1, value="ADC_REFH_BTM_V_CAL2")
    ws.cell(row=8, column=col_start_at1+6+7+8+1, value="ADC_BTM_VHGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+7+7+8+1, value="ADC_REFH_BTM_V_CAL3")
    ws.cell(row=8, column=col_start_at1+8+7+8+1, value="ADC_BTM_VHGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+9+7+8+1, value="ADC_REFH_BTM_V_CAL4")
    ws.cell(row=8, column=col_start_at1+10+7+8+1, value="ADC_BTM_VHGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+11+7+8+1, value="ADC_REFH_BTM_V_CAL5")
    ws.cell(row=9, column=col_start_at1+0+7+8+1, value="V")
    ws.cell(row=row_usl, column=col_start_at1+0+7+8+1, value=ADC_REFH_BTM_V_USL)
    ws.cell(row=row_lsl, column=col_start_at1+0+7+8+1, value=ADC_REFH_BTM_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_ADC_REFH_BTM_VHGEN = 4
    global ADC_REFH_BTM_VHGEN
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BE, 0x10)
    time.sleep(1/1000)
    btm_adc_refh_voltage = measure_adc_refh_btm_voltage()
    ws.cell(row=row_start, column=col_start_at1+16, value=btm_adc_refh_voltage) # col = 126
    print(f"ADC_REFH_BTM Voltage{cnt}:{btm_adc_refh_voltage}V")
    
    if ADC_REFH_BTM_V_LSL <= btm_adc_refh_voltage <= ADC_REFH_BTM_V_USL:    
        addr05BE = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE))
        addr05BE = int(addr05BE, 16)
        ADC_REFH_BTM_VHGEN = addr05BE
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start_at1+16+1, value=INDICATOR)
        while(cnt < 5):
            if btm_adc_refh_voltage < ADC_REFH_BTM_V_LSL:
                if cnt == 0:
                    DELTA_BTM_VHGEN = 15
                elif cnt >= 1:
                    DELTA_BTM_VHGEN = 0
                    DELTA_BTM_VHGEN += (2**(N_ADC_REFH_BTM_VHGEN-cnt)) #8, 4, 2, 1

                addr05BE = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE))
                addr05BE = int(addr05BE, 16)
                ADC_REFH_BTM_VHGEN = addr05BE
                ADC_REFH_BTM_VHGEN += DELTA_BTM_VHGEN                
                ws.cell(row=row_start, column=(col_start_at1+16) + (2 * cnt) + 2, value=ADC_REFH_BTM_VHGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_BTM_VHGEN:{DELTA_BTM_VHGEN}, ADC_REFH_BTM_VHGEN:{ADC_REFH_BTM_VHGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BE, ADC_REFH_BTM_VHGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE)))
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+16) + (2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFH_BTM Voltage{cnt}:{voltage_read}V")
                
                cnt += 1    
                
                btm_adc_refh_voltage = voltage_read                
                if ADC_REFH_BTM_V_LSL <= voltage_read <= ADC_REFH_BTM_V_USL:
                    break

            elif btm_adc_refh_voltage > ADC_REFH_BTM_V_LSL:
                DELTA_BTM_VHGEN = 0
                DELTA_BTM_VHGEN -= (2**(N_ADC_REFH_BTM_VHGEN-cnt)) # 16, 8, 4, 2, 1
                addr05BE = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE))
                addr05BE = int(addr05BE, 16)
                ADC_REFH_BTM_VHGEN = addr05BE
                ADC_REFH_BTM_VHGEN += DELTA_BTM_VHGEN

                ws.cell(row=row_start, column=(col_start_at1+16) + (2 * cnt) + 2, value=ADC_REFH_BTM_VHGEN)
                print(f"UP, cnt:{cnt}, DELTA_BTM_VHGEN:{DELTA_BTM_VHGEN}, ADC_REFH_BTM_VHGEN:{ADC_REFH_BTM_VHGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BE, ADC_REFH_BTM_VHGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE)))                            
                
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+16)+(2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFH_BTM Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                btm_adc_refh_voltage = voltage_read
                if ADC_REFH_BTM_V_LSL <= voltage_read <= ADC_REFH_BTM_V_USL:
                    break
    
    print(f"ADC_REFH_BTM_V_CAL_CNT:{cnt}")
    print("-----END ADC_REFH BTM VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_adc_refl_top_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START ADC_REFL TOP VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start+2+0+7+8+1+1+10, value="ADC_REFL_TOP_VOLT")
    ws.cell(row=8, column=col_start+2+1+7+8+1+1+10, value="ADC_REFL_TOP_V_INDCATOR")
    ws.cell(row=8, column=col_start+2+2+7+8+1+1+10, value="ADC_TOP_VLGEN<4:0>")
    ws.cell(row=8, column=col_start+2+3+7+8+1+1+10, value="ADC_REFL_TOP_V_CAL1")
    ws.cell(row=8, column=col_start+2+4+7+8+1+1+10, value="ADC_TOP_VLGEN<4:0>")
    ws.cell(row=8, column=col_start+2+5+7+8+1+1+10, value="ADC_REFL_TOP_V_CAL2")
    ws.cell(row=8, column=col_start+2+6+7+8+1+1+10, value="ADC_TOP_VLGEN<4:0>")
    ws.cell(row=8, column=col_start+2+7+7+8+1+1+10, value="ADC_REFL_TOP_V_CAL3")
    ws.cell(row=8, column=col_start+2+8+7+8+1+1+10, value="ADC_TOP_VLGEN<4:0>")
    ws.cell(row=8, column=col_start+2+9+7+8+1+1+10, value="ADC_REFL_TOP_V_CAL4")
    ws.cell(row=8, column=col_start+2+10+7+8+1+1+10, value="ADC_TOP_VLGEN<4:0>")
    ws.cell(row=8, column=col_start+2+11+7+8+1+1+10, value="ADC_REFL_TOP_V_CAL5")
    ws.cell(row=9, column=col_start+2+0+7+8+1+1+10, value="V")
    ws.cell(row=row_usl, column=col_start+2+7+8+1+1+10, value=ADC_REFL_TOP_V_USL)
    ws.cell(row=row_lsl, column=col_start+2+7+8+1+1+10, value=ADC_REFL_TOP_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_ADC_REFL_TOP_VLGEN = 4
    global ADC_REFL_TOP_VLGEN
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BF, 0x10)
    time.sleep(1/1000)
    top_adc_refl_voltage = measure_adc_refl_top_voltage()
    ws.cell(row=row_start, column=col_start+2+7+8+1+1+10, value=top_adc_refl_voltage) 
    print(f"ADC_REFL_TOP Voltage{cnt}:{top_adc_refl_voltage}V")
    
    if ADC_REFL_TOP_V_LSL <= top_adc_refl_voltage <= ADC_REFL_TOP_V_USL:    
        addr05BF = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF))
        addr05BF = int(addr05BF, 16)
        ADC_REFL_TOP_VLGEN = addr05BF
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start+2+7+8+1+1+10+1, value=INDICATOR)
        while(cnt < 5):
            if top_adc_refl_voltage < ADC_REFL_TOP_V_LSL:
                if cnt == 0:
                    DELTA_TOP_VLGEN = 15
                elif cnt >= 1:
                    DELTA_TOP_VLGEN = 0
                    DELTA_TOP_VLGEN += (2**(N_ADC_REFL_TOP_VLGEN-cnt)) #8, 4, 2, 1

                addr05BF = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF))
                addr05BF = int(addr05BF, 16)
                ADC_REFL_TOP_VLGEN = addr05BF
                ADC_REFL_TOP_VLGEN += DELTA_TOP_VLGEN
                ws.cell(row=row_start, column=(col_start+2+7+8+1+1+10) + (2 * cnt) + 2, value=ADC_REFL_TOP_VLGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_TOP_VLGEN:{DELTA_TOP_VLGEN}, ADC_REFL_TOP_VLGEN:{ADC_REFL_TOP_VLGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BF, ADC_REFL_TOP_VLGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF)))
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+2+7+8+1+1+10)+(2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFL_TOP Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                top_adc_refl_voltage = voltage_read                
                if ADC_REFL_TOP_V_LSL <= voltage_read <= ADC_REFL_TOP_V_USL:
                    break
            elif top_adc_refl_voltage > ADC_REFL_TOP_V_LSL:
                DELTA_TOP_VLGEN = 0
                DELTA_TOP_VLGEN -= (2**(N_ADC_REFL_TOP_VLGEN-cnt)) # 8, 4, 2, 1
                addr05BF = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF))
                addr05BF = int(addr05BF, 16)
                ADC_REFL_TOP_VLGEN = addr05BF
                ADC_REFL_TOP_VLGEN += DELTA_TOP_VLGEN
                      
                ws.cell(row=row_start, column=(col_start+2+7+8+1+1+10) + (2 * cnt) + 2, value=ADC_REFL_TOP_VLGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_TOP_VLGEN:{DELTA_TOP_VLGEN}, ADC_REFL_TOP_VLGEN:{ADC_REFL_TOP_VLGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05BF, ADC_REFL_TOP_VLGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF)))

                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start+2+7+8+1+1+10)+(2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFL_TOP Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                top_adc_refl_voltage = voltage_read
                if ADC_REFL_TOP_V_LSL <= voltage_read <= ADC_REFL_TOP_V_USL:
                    break
    
    print(f"ADC_REFL_TOP_V_CAL_CNT:{cnt}")
    print("-----END ADC_REFL TOP VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def calibration_adc_refl_btm_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START ADC_REFL BTM VOLTAGE CALIBRATION-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+0+7+8+1+1+11, value="ADC_REFL_BTM_VOLT")
    ws.cell(row=8, column=col_start_at1+1+7+8+1+1+11, value="ADC_REFL_BTM_V_INDCATOR")
    ws.cell(row=8, column=col_start_at1+2+7+8+1+1+11, value="ADC_BTM_VLGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+3+7+8+1+1+11, value="ADC_REFL_BTM_V_CAL1")
    ws.cell(row=8, column=col_start_at1+4+7+8+1+1+11, value="ADC_BTM_VLGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+5+7+8+1+1+11, value="ADC_REFL_BTM_V_CAL2")
    ws.cell(row=8, column=col_start_at1+6+7+8+1+1+11, value="ADC_BTM_VLGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+7+7+8+1+1+11, value="ADC_REFL_BTM_V_CAL3")
    ws.cell(row=8, column=col_start_at1+8+7+8+1+1+11, value="ADC_BTM_VLGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+9+7+8+1+1+11, value="ADC_REFL_BTM_V_CAL4")
    ws.cell(row=8, column=col_start_at1+10+7+8+1+1+11, value="ADC_BTM_VLGEN<4:0>")
    ws.cell(row=8, column=col_start_at1+11+7+8+1+1+11, value="ADC_REFL_BTM_V_CAL5")
    ws.cell(row=9, column=col_start_at1+0+7+8+1+1+11, value="V")
    ws.cell(row=row_usl, column=col_start_at1+0+7+8+1+1+11, value=ADC_REFL_BTM_V_USL)
    ws.cell(row=row_lsl, column=col_start_at1+0+7+8+1+1+11, value=ADC_REFL_BTM_V_LSL)

    cnt = 0
    INDICATOR = 0
    N_ADC_REFL_BTM_VLGEN = 4
    global ADC_REFL_BTM_VLGEN
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C0, 0x10)
    time.sleep(1/1000)
    btm_adc_refl_voltage = measure_adc_refl_btm_voltage()
    ws.cell(row=row_start, column=col_start_at1+28, value=btm_adc_refl_voltage)
    print(f"ADC_REFL_BTM Voltage{cnt}:{btm_adc_refl_voltage}V")
    
    if ADC_REFL_BTM_V_LSL <= btm_adc_refl_voltage <= ADC_REFL_BTM_V_USL:    
        addr05C0 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0))
        addr05C0 = int(addr05C0, 16)
        ADC_REFL_BTM_VLGEN = addr05C0
        print("CNT:",cnt, "NO CALIBRATION")
    else:
        INDICATOR += 1
        ws.cell(row=(row_start+0), column=col_start_at1+28+1, value=INDICATOR) 
        while(cnt < 5):
            if btm_adc_refl_voltage < ADC_REFL_BTM_V_LSL:
                if cnt == 0:
                    DELTA_BTM_VLGEN = 15
                elif cnt >= 1:
                    DELTA_BTM_VLGEN = 0
                    DELTA_BTM_VLGEN += (2**(N_ADC_REFL_BTM_VLGEN-cnt))
                
                addr05C0 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0))
                addr05C0 = int(addr05C0, 16)
                ADC_REFL_BTM_VLGEN = addr05C0
                ADC_REFL_BTM_VLGEN += DELTA_BTM_VLGEN
                ws.cell(row=row_start, column=(col_start_at1+28) + (2 * cnt) + 2, value=ADC_REFL_BTM_VLGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_BTM_VLGEN:{DELTA_BTM_VLGEN}, ADC_REFL_BTM_VLGEN:{ADC_REFL_BTM_VLGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05C0, ADC_REFL_BTM_VLGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0)))
                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+28)+(2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFL_BTM Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                btm_adc_refl_voltage = voltage_read                
                if ADC_REFL_BTM_V_LSL <= voltage_read <= ADC_REFL_BTM_V_USL:
                    break
            elif btm_adc_refl_voltage > ADC_REFL_BTM_V_LSL:
                DELTA_BTM_VLGEN = 0
                DELTA_BTM_VLGEN -= (2**(N_ADC_REFL_BTM_VLGEN-cnt)) # 8, 4, 2, 1
                addr05C0 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0))
                addr05C0 = int(addr05C0, 16)
                ADC_REFL_BTM_VLGEN = addr05C0
                ADC_REFL_BTM_VLGEN += DELTA_BTM_VLGEN
                      
                ws.cell(row=row_start, column=(col_start_at1+28) + (2 * cnt) + 2, value=ADC_REFL_BTM_VLGEN)
                print(f"DOWN, cnt:{cnt}, DELTA_BTM_VLGEN:{DELTA_BTM_VLGEN}, ADC_REFL_BTM_VLGEN:{ADC_REFL_BTM_VLGEN}")
                evb.i2c0_reg16_write(SlaveAddr_0, 0x05C0, ADC_REFL_BTM_VLGEN)
                print(hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0)))

                smu.gpib.write(":SOUR:FUNC CURR")
                time.sleep(0.1)
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = float(voltage_read)
                ws.cell(row=row_start, column=(col_start_at1+28)+(2 * cnt) + 3, value=voltage_read)
                print(f"ADC_REFL_BTM Voltage{cnt}:{voltage_read}V")
                
                cnt += 1
                
                btm_adc_refl_voltage = voltage_read
                if ADC_REFL_BTM_V_LSL <= voltage_read <= ADC_REFL_BTM_V_USL:
                    break
    
    print(f"ADC_REFL_BTM_V_CAL_CNT:{cnt}")
    print("-----END ADC_REFL BTM VOLTAGE CALIBRATION-----")
    wb.save(save_path)
    wb.close()

def measure_mipirx_bias_voltage():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START MIPIRX BIAS VOLTAGE MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    mipirx_dict = {"MIPIRX_VILCD":(0x18, (MIPIRX_VILCD_LSL, MIPIRX_VILCD_USL)),\
                    "MIPIRX_VIHCD":(0x20, (MIPIRX_VIHCD_LSL, MIPIRX_VIHCD_USL)),\
                    "MIPIRX_VULPS":(0x28, (MIPIRX_VULPS_LSL, MIPIRX_VULPS_USL)),\
                    "MIPIRX_VIL":(0x30, (MIPIRX_VIL_LSL, MIPIRX_VIL_USL)),\
                    "MIPIRX_VIH":(0x38, (MIPIRX_VIH_LSL, MIPIRX_VIH_USL))}    
    for idx, items in enumerate(mipirx_dict.items()):
        ws.cell(row=8, column=col_start_at1+87+idx, value=items[0])
        ws.cell(row=9, column=col_start_at1+87+idx, value="V")
        ws.cell(row=row_usl, column=col_start_at1+87+idx, value=items[-1][-1][-1])
        ws.cell(row=row_lsl, column=col_start_at1+87+idx, value=items[-1][-1][0])
        evb.i2c0_reg16_write(SlaveAddr_0, 0x002e, items[-1][0])
        time.sleep(1/1000)
        smu.gpib.write(":SOUR:FUNC CURR")
        time.sleep(1/10)
        voltage_val = smu.measure_voltage()
        voltage_read, _, _, _, _ = voltage_val.split(",")
        voltage_read = float(voltage_read)
        ws.cell(row=row_start, column=col_start_at1+87+idx, value=voltage_read)
        print(f"{items[0]}_Voltage:{voltage_read}V")

    print("-----END MIPIRX BIAS VOLTAGE MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_bist_functions():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START BIST FUNCTIONS MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    # ws.cell(row=8, column=col_start_at1+88, value="ROW_STUCK_LOW_DATA")
    ws.cell(row=8, column=col_start_at1+92, value="ROW_STUCK_LOW_DATA") # 202
    ws.cell(row=row_usl, column=col_start_at1+92, value=3) # pass value(dec)
    ws.cell(row=row_lsl, column=col_start_at1+92, value=3) # pass value(dec)
    ws.cell(row=8, column=col_start_at1+93, value="ROW_STUCK_LOW_RESULT")
    ws.cell(row=9, column=col_start_at1+93, value="P/F")    
    ws.cell(row=8, column=col_start_at1+94, value="ROW_STUCK_HIGH_DATA")
    ws.cell(row=row_usl, column=col_start_at1+94, value=48) # pass value(dec)
    ws.cell(row=row_lsl, column=col_start_at1+94, value=48) # pass value(dec)
    ws.cell(row=8, column=col_start_at1+95, value="ROW_STUCK_HIGH_RESULT")
    ws.cell(row=9, column=col_start_at1+95, value="P/F")    
    ### row bist ###

    row_stuck0_pass = 3 # 0x03
    row_stuck1_pass = 48 # 0x30

    evb.i2c0_reg16_write(SlaveAddr_0, 0x00FE, 0x0F) # Row check enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00FA, 0x77) # scan manual setting(stuck-at-low mode)
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00FB, 0x11) # em manual setting(stuck-at-low mode)
    time.sleep(1/1000)
    result_row = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x00FF)) # 0x03 -> PASS
    result_row = int(result_row, 16)
    time.sleep(1/1000)
    print(f"Row_stuck_low_result :{result_row}(PASS at {row_stuck0_pass})")
    if result_row == row_stuck0_pass:
        ws.cell(row=12, column=col_start_at1+92, value=result_row)
        ws.cell(row=12, column=col_start_at1+93, value="PASS")
    else:
        ws.cell(row=12, column=col_start_at1+92, value=result_row)
        ws.cell(row=12, column=col_start_at1+93, value="FAIL")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x00FA, 0x07) # scan manual setting(stuck-at-high mode)
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00FB, 0x01) # em manual setting(stuck-at-high mode)
    time.sleep(1/1000)
    result_row = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x00FF)) # 0x30 -> PASS
    result_row = int(result_row, 16)
    time.sleep(1/1000)
    print(f"Row_stuck_high_result :{result_row}(PASS at {row_stuck1_pass})")
    if result_row == row_stuck1_pass:
        ws.cell(row=12, column=col_start_at1+94, value=result_row)
        ws.cell(row=12, column=col_start_at1+95, value="PASS")
    else:
        ws.cell(row=12, column=col_start_at1+94, value=result_row)
        ws.cell(row=12, column=col_start_at1+95, value="PASS")
    
    ### column bist ###
    data_bit = 12
    data_lst = [2**data for data in range(0, data_bit+1)]
    column_pass = 16
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00DF, 0x01) # column check enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00DA, 0x01) # col_bist_ref_h_sel

    for idx, data in enumerate(data_lst):
        if data == 4096:
            data -= 1
        ws.cell(row=8, column=col_start_at1+96+idx, value=f"COLUMN_BIST_DATA{data}")
        ws.cell(row=9, column=col_start_at1+96+idx, value="P/F")

        evb.i2c0_reg16_write(SlaveAddr_0, 0x00D6, (data>>8) + 128) # col_data<15:8>
        time.sleep(1/1000)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00D7, data) # col_data<7:0>
        time.sleep(1/1000)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00DB, data>>8) # ref_vh<15:8>
        time.sleep(1/1000)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00DC, data) # ref_vh<7:0>
        time.sleep(1/1000)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00E0, 0x01) # col_bist trigger
        time.sleep(1/1000)
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00E0, 0x00) # col_bist trigger
        time.sleep(1/1000)
        result_col = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x00E2)) # col_bist result
        result_col = int(result_col, 16)
        time.sleep(1/1000)
        print(f"Column_bist_result :{result_col}(PASS at {column_pass})")
        if result_col == column_pass:
            ws.cell(row=12, column=col_start_at1+96+idx, value=result_col)
            ws.cell(row=12, column=col_start_at1+96+idx, value="PASS")
        else:
            ws.cell(row=12, column=col_start_at1+96+idx, value=result_col)
            ws.cell(row=12, column=col_start_at1+96+idx, value="FAIL")

    print("-----END BIST FUNCTIONS MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_adc_external_sweep():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START ADC EXTERNAL VOLTAGE SWEEP MEAUSUREMENT-----")
    wb, ws = option_create_sheet()

    ws.cell(row=8, column=col_start_at1+109, value="ADC_EXTERNAL_INPUT_VOLT")
    ws.cell(row=9, column=col_start_at1+109, value="V")    

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002E, 0x05) # BTM ADC external input set(AT1)
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002F, 0x28) # TOP ADC external input set(AT0)
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x01) # ADC_TOP_SOC enable
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BC, 0x0A) # TOP,BTM ADC external input set
    time.sleep(1/1000)
    evb.i2c0_reg16_write(SlaveAddr_0, 0x05D1, 0xF0) # TEMP_FILT_MODE:avg
    time.sleep(1/100)
    
    smu.gpib.write(":SOUR:FUNC VOLT") # SMU voltage source setting
    smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert FRONT Terminal
    smu.enable_output()
    time.sleep(1)
    
    row_sweep = 12
    adc_refl = 0.5
    adc_refh = 3.5
    adc_bit = 1024 # 10 bit
    step = (adc_refh - adc_refl) / adc_bit # 1024
    for n in range(0, adc_bit, 1): # 0 ~ 1023
        ext_volt = adc_refl + (n*step)    
        if n == adc_bit-1:
            ext_volt = adc_refh
        ws.cell(row=row_sweep, column=col_start_at1+109, value=ext_volt)
        smu.set_voltage(ext_volt) # SMU voltage source volt setting
        time.sleep(0.1)

        temp_cnt = 11
        temp_top_list = []
        for i in range(1, temp_cnt+1, 1): # 1 ~ 11
            if i == 1:
                ws.cell(row=9, column=col_start_at1+109+i, value="CODE(DEC)")
            if i == temp_cnt:
                ws.cell(row=8, column=col_start_at1+109+i, value=f"TOP_TEMP_READ_AF(avg)")
                ws.cell(row=row_sweep, column=col_start_at1+109+i, value=f"{int(sum(temp_top_list)/len(temp_top_list))}") 
                break

            temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C5) # TEMPERATURE_LT<9:2>
            time.sleep(1/1000)
            temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C6) # TEMPERATURE_LT<1:0>
            temp_10bit = (temp_msb8<<2) + temp_lsb2
            temp_10bit = int(temp_10bit)
            time.sleep(1/1000)
            temp_top_list.append(temp_10bit)

            ws.cell(row=8, column=col_start_at1+109+i, value=f"TOP_TEMP_READ{i}_AF")
            ws.cell(row=row_sweep, column=col_start_at1+109+i, value=f"{temp_10bit}")       
            print(f"TOP_TEMPCODE_AF{i} : {temp_10bit}")
        
        row_sweep += 1

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C1, 0x02) # ADC_BTM_SOC enable
    time.sleep(1/1000)
    smu.gpib.write(":ROUTe:TERMinals REAR") # Convert REAR Terminal
    smu.enable_output()
    time.sleep(1)

    row_sweep = 12
    for n in range(0, adc_bit, 1):
        ext_volt = adc_refl + (n*step)   
        if n == adc_bit-1:
            ext_volt = adc_refh 
        smu.set_voltage(ext_volt) # SMU voltage source volt setting
        time.sleep(0.1) # 1

        temp_cnt = 11
        temp_btm_list = []
        for i in range(temp_cnt+1, (2*temp_cnt)+1, 1): # 12 ~ 22
            if i == temp_cnt+1:
                ws.cell(row=9, column=col_start_at1+109+i, value="CODE(DEC)")
            if i == (2*temp_cnt):
                ws.cell(row=8, column=col_start_at1+109+i, value=f"BTM_TEMP_READ_AF(avg)")
                ws.cell(row=row_sweep, column=col_start_at1+109+i, value=f"{int(sum(temp_btm_list)/len(temp_btm_list))}") 
                break

            temp_msb8 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05C9) # TEMPERATURE_LB<9:2>
            time.sleep(1/1000)
            temp_lsb2 = evb.i2c0_reg16_read(SlaveAddr_0, 0x05CA) # TEMPERATURE_LB<1:0>
            temp_10bit = (temp_msb8<<2) + temp_lsb2
            temp_10bit = int(temp_10bit)
            time.sleep(1/1000)
            temp_btm_list.append(temp_10bit)

            ws.cell(row=8, column=col_start_at1+109+i, value=f"BTM_TEMP_READ{i-11}_AF")
            ws.cell(row=row_sweep, column=col_start_at1+109+i, value=f"{temp_10bit}")       
            print(f"BTM_TEMPCODE_AF{i-11} : {temp_10bit}")
        
        row_sweep += 1
    print("-----END ADC EXTERNAL VOLTAGE SWEEP MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_column_data_sweep():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START COLUMN DATA SWEEP MEAUSUREMENT-----")
    wb, ws = option_create_sheet()
    
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f0, 0x01) # col_test_en
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0426, 0x10) # agamma_fix_en
    time.sleep(1/1000) # 1ms

    evb.i2c0_reg16_write(SlaveAddr_0, 0x002F, 0x10) # column output voltage
    time.sleep(1/1000) # 1ms
    smu.gpib.write(":SOUR:FUNC CURR") # SMU current source setting
    time.sleep(1/10) # 0.1s

    N_start = 0
    N_finish = 4095 # data bit
    N_step = 1
    col_test_addr = [0, 187, 375]
    col_datasweep = 242
    ws.cell(row=8, column=col_datasweep, value="COL_DATA")
    ws.cell(row=9, column=col_datasweep, value="CODE(DEC)") 
    for col_test in col_test_addr:
        ws.cell(row=8, column=col_datasweep+1, value=f"COL_ADDR{col_test}")
        ws.cell(row=9, column=col_datasweep+1, value="V")

        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f1, col_test>>8) # col_test_addr<8>
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f2, col_test) # col_test_addr<7:0>
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x01) # col_test_trig
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x00) # col_test_trig
        time.sleep(1/1000) # 1ms

        row_datasweep = 12
        for data in range(N_start, N_finish+1, N_step):
            # if data == N_finish + 1:
            #     data = N_finish
            evb.i2c0_reg16_write(SlaveAddr_0, 0x00d6, (data>>8) + 128) # col_man_en & col_data_man<11:8>
            time.sleep(1/1000) # 1ms
            evb.i2c0_reg16_write(SlaveAddr_0, 0x00d7, data) # col_data_man<7:0>
            time.sleep(1/1000) # 1ms
            ws.cell(row=row_datasweep, column=242, value=data) # col = col_datasweep
            voltage_val = smu.measure_voltage()
            voltage_read, _, _, _, _ = voltage_val.split(",")
            voltage_read = float(voltage_read)
            ws.cell(row=row_datasweep, column=col_datasweep+1, value=voltage_read)
            print(f"{col_test}_test_column_data({data})_output_voltage : {voltage_read}V")
            row_datasweep += 1

        col_datasweep += 1

    print("-----END COLUMN DATA SWEEP MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def measure_gamma_tapcon_sweep():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("-----START GAMMA TAPCON SWEEP MEAUSUREMENT-----")
    wb, ws = option_create_sheet()
    
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f0, 0x01) # col_test_en
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x0426, 0x10) # agamma_fix_en
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x002F, 0x10) # column output voltage
    time.sleep(1/1000) # 1ms
    smu.gpib.write(":SOUR:FUNC CURR") # SMU current source setting
    time.sleep(1/10) # 0.1s

    N_start = 0
    N_finish = 4095 # data bit
    N_step = 1
    col_test_addr = 187
    col_data_lst = [16, 32, 48, 64, 80, 96, 112, 256, 416, 576, 736, 896, 1056, 1216, 1376, 1536,\
                    1696, 1856, 2016, 2176, 2336, 2496, 2656, 2816, 2976, 3136, 3296, 3456, 3616,\
                    3776, 3936, 4095] # GMATAP DATA
    col_datasweep = 234 + 6
    bit_tapcon = 1024 # 10bit
    step_tapcon = 1
    ws.cell(row=8, column=col_datasweep, value="TAPCON_DATA")
    ws.cell(row=9, column=col_datasweep, value="CODE(DEC)") # 240
    ws.cell(row=8, column=col_datasweep+1, value=f"COL_OUTPUT{col_test_addr}")
    ws.cell(row=9, column=col_datasweep+1, value="V") # 241

    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f1, col_test_addr>>8) # col_test_addr<8>
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f2, col_test_addr) # col_test_addr<7:0>
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x01) # col_test_trig
    time.sleep(1/1000) # 1ms
    evb.i2c0_reg16_write(SlaveAddr_0, 0x00f3, 0x00) # col_test_trig
    time.sleep(1/1000) # 1ms
    # for data in range(N_start, N_finish+2, N_step):
        # if data == N_finish + 1:
            # data = N_finish
    for idx, data in enumerate(col_data_lst): # idx : 0 ~ 31
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00d6, (data>>8) + 128) # col_man_en & col_data_man<11:8>
        time.sleep(1/1000) # 1ms
        evb.i2c0_reg16_write(SlaveAddr_0, 0x00d7, data) # col_data_man<7:0>
        ws.cell(row=8, column=col_datasweep+1, value=f"COLUMN_OUTPUT_{data}")
        row_datasweep = 12
        for data_tapcon in range(0, bit_tapcon, step_tapcon): # 0 ~ 1023 sweep
            tapcon_dict = {f"RED_PING_TAPCON{idx}":(1066+(2*idx), 1067+(2*idx)),\
                            f"GRU_PING_TAPCON{idx}":(1130+(2*idx), 1131+(2*idx)),\
                            f"BLU_PING_TAPCON{idx}":(1194+(2*idx), 1195+(2*idx)),\
                            }
            for items in tapcon_dict.items():                    
                evb.i2c0_reg16_write(SlaveAddr_0, items[-1][0], data_tapcon>>8) # rgb_ping_tapcon<9:8>
                time.sleep(1/1000) # 1ms
                evb.i2c0_reg16_write(SlaveAddr_0, items[-1][-1], data_tapcon) # rgb_ping_tapcon<7:0>
                time.sleep(1/10) # 100ms             
            voltage_lst = []
            cnt = 0
            while(cnt < 4):
                cnt += 1
                voltage_val = smu.measure_voltage()
                voltage_read, _, _, _, _ = voltage_val.split(",")
                voltage_read = round(float(voltage_read), 6)
                voltage_lst.append(voltage_read)
                print(f"{items[0][9:]}:{data_tapcon}, {col_test_addr}_test_column_data({data})_output_voltage : {voltage_read}V")
            voltage_avg = sum(voltage_lst) / len(voltage_lst)
            voltage_avg = round(float(voltage_avg), 6)
            ws.cell(row=row_datasweep, column=240, value=data_tapcon)
            ws.cell(row=row_datasweep, column=col_datasweep+1, value=voltage_avg)
            print(f"{items[0][9:]}:{data_tapcon}, {col_test_addr}_test_column_data({data})_output_voltage(avg) : {voltage_avg}V")
            row_datasweep += 1
        col_datasweep += 1
        wb.save(save_path)
    print("-----END GAMMA TAPCON SWEEP MEAUSUREMENT-----")
    wb.save(save_path)
    wb.close()

def read_calibration_values():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("START READ CALIBRATION VALUES")

    addr0012 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012))
    time.sleep(1/1000)
    addr0012 = int(addr0012, 16)
    global BIAS_TOP_CON
    BIAS_TOP_CON = addr0012 

    addr0013 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013))
    time.sleep(1/1000)
    addr0013 = int(addr0013, 16)
    global BIAS_BTM_CON
    BIAS_BTM_CON = addr0013

    addr05BD = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD))
    time.sleep(1/1000)
    addr05BD = int(addr05BD, 16)
    global ADC_REFH_TOP_CON
    ADC_REFH_TOP_CON = addr05BD

    addr05BE = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE))
    time.sleep(1/1000)
    addr05BE = int(addr05BE, 16)
    global ADC_REFH_BTM_CON
    ADC_REFH_BTM_CON = addr05BE

    addr05BF = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF))
    time.sleep(1/1000)
    addr05BF = int(addr05BF, 16)
    global ADC_REFL_TOP_CON
    ADC_REFL_TOP_CON = addr05BF

    addr05C0 = hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0))
    time.sleep(1/1000)
    addr05C0 = int(addr05C0, 16)
    global ADC_REFL_BTM_CON
    ADC_REFL_BTM_CON = addr05C0
    
    cal_list = [("BIAS_TOP_CON",BIAS_TOP_CON),\
                ("BIAS_BTM_CON",BIAS_BTM_CON),\
                ("ADC_REFH_TOP_CON",ADC_REFH_TOP_CON),\
                ("ADC_REFL_TOP_CON",ADC_REFL_TOP_CON),\
                ("ADC_REFH_BTM_CON",ADC_REFH_BTM_CON),\
                ("ADC_REFL_BTM_CON",ADC_REFL_BTM_CON)]
    for cal_val in cal_list:
        print(f"{cal_val[0]} : {hex(cal_val[-1])}")

    print("END READ CALIBRATION VALUES")

def write_calibration_values():
    global func_name
    global total_progress
    func_name = inspect.currentframe().f_code.co_name
    total_progress += 1

    print("START WRITE CALIBRATION VALUES")

    evb.i2c0_reg16_write(SlaveAddr_0, 0x0012, BIAS_TOP_CON) # TOP HVBGR VCON,RCON
    time.sleep(1/1000)
    print("BIAS_TOP_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0012)))

    evb.i2c0_reg16_write(SlaveAddr_0, 0x0013, BIAS_BTM_CON) # BTM HVBGR VCON,RCON
    time.sleep(1/1000)
    print("BIAS_BTM_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x0013)))

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BD, ADC_REFH_TOP_CON)
    time.sleep(1/1000)
    print("ADC_REFH_TOP_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BD)))

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BE, ADC_REFH_BTM_CON)
    time.sleep(1/1000)
    print("ADC_REFH_BTM_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BE)))

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05BF, ADC_REFL_TOP_CON)
    time.sleep(1/1000)
    print("ADC_REFL_TOP_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05BF)))

    evb.i2c0_reg16_write(SlaveAddr_0, 0x05C0, ADC_REFL_BTM_CON)
    time.sleep(1/1000)
    print("ADC_REFL_BTM_CON:", hex(evb.i2c0_reg16_read(SlaveAddr_0, 0x05C0)))

    print("END WRITE CALIBRATION VALUES")


########## MEASUREMENT SETTING ##########
smu = KEITHLEY2401()
smu.set_verbose(True)
# smu.open()
smu.open("01211918") # KEITHLEY GPIB Serial number
# smu.enable_output()

evb = EVB()
evb.open()
evb.i2c0_change_port(0)

chamber = su.SU241()
chamber.open("01209782") # CHAMBER GPIB Serial number
# chamber.open() # CHAMBER GPIB
# chamber.open("01211918") # KEITHLEY GPIB

########## TEST LIMIT SETTING ##########
HVBGR_TOP_V_USL = 1.227 # Unit : V
HVBGR_TOP_V_LSL = 1.215 # Unit : V
HVBGR_TOP_I_USL = 5.19E-06 # Unit : A
HVBGR_TOP_I_LSL = 4.95E-06 # Unit : A
HVBGR_BTM_V_USL = 1.227 # Unit : V
HVBGR_BTM_V_LSL = 1.215 # Unit : V
HVBGR_BTM_I_USL = 5.19E-06 # Unit : A
HVBGR_BTM_I_LSL = 4.95E-06 # Unit : A
LDO33_V_USL = 3.45 # Unit : V
LDO33_V_LSL = 3.15 # Unit : V
ADC_REFH_TOP_V_USL = 3.520 # Unit : V
ADC_REFH_TOP_V_LSL = 3.488 # Unit : V
ADC_REFL_TOP_V_USL = 0.509 # Unit : V
ADC_REFL_TOP_V_LSL = 0.503 # Unit : V
ADC_REFH_BTM_V_USL = 3.520 # Unit : V
ADC_REFH_BTM_V_LSL = 3.488 # Unit : V
ADC_REFL_BTM_V_USL = 0.509 # Unit : V
ADC_REFL_BTM_V_LSL = 0.503 # Unit : V
TEMP_CODE_LT_USL = 542 # Unit : decimal @50'C
TEMP_CODE_LT_LSL = 523 # Unit : decimal @50'C
TEMP_CODE_LB_USL = 542 # Unit : decimal @50'C
TEMP_CODE_LB_LSL = 523 # Unit : decimal @50'C
TEMP_CODE_RT_USL = 535 # Unit : decimal @50'C
TEMP_CODE_RT_LSL = 515 # Unit : decimal @50'C
TEMP_CODE_RB_USL = 535 # Unit : decimal @50'C
TEMP_CODE_RB_LSL = 515 # Unit : decimal @50'C
MAX_SUB_PIXEL_I_USL = 1.8E-06 # A
MAX_SUB_PIXEL_I_LSL = 1.3E-06 # A
MIPIRX_VILCD_USL = "N/A" # Unit : V
MIPIRX_VILCD_LSL = 0.200 # Unit : V
MIPIRX_VIHCD_USL = 0.450 # Unit : V
MIPIRX_VIHCD_LSL = "N/A" # Unit : V
MIPIRX_VULPS_USL = "N/A" # Unit : V
MIPIRX_VULPS_LSL = 0.300 # Unit : V
MIPIRX_VIL_USL = "N/A" # Unit : V
MIPIRX_VIL_LSL = 0.550 # Unit : V
MIPIRX_VIH_USL = 0.880 # Unit : V
MIPIRX_VIH_LSL = "N/A" # Unit : V

########## DATE & TIME SETTING ##########
now = datetime.now()

########## GUI SETTING ##########
root=tkinter.Tk()
root.title('TEST STATUS INFORMATION')
root.geometry('450x100+400+400')
var=tkinter.DoubleVar()
progressbar=tkinter.ttk.Progressbar(root, variable=var, length=300, mode="determinate")
progressbar.pack(fill="x", padx=5, pady=5)
label=tkinter.Label(root ,text='')
label2=tkinter.Label(root ,text='')
label3=tkinter.Label(root ,text='')
label.pack()
label2.pack()
label3.pack()

########## TEST INFO SETTING ##########
evboard_num = input('\nevboard_number : ') # Enter EVboard number
chip_num = input('\ntest_sample_number : ') # Enter DUT number
tester_name = input('\ntester_name : ') # Enter Tester name
col2 = 6 # I2C log file data start column 
col_start = col2 + 14 # Measurement log data start column(20)
col_start_at1 = col_start + 90 # 106 -> 110
col_start_all_measure = 248 # after calibration
row_start = 12 # log data start row

########## FILES PATH SETTING ##########
regmap_path = '/Users/rlee/Desktop/Ronlee/RegMap_RDP180XP_REV0_241025_RAON_copy.xlsx' # RDP180XP register map path
save_path = f"/Users/rlee/Desktop/Ronlee/RDP180XP_electrical_test_results_{chip_num}.xlsx" # If you have an existing logfile_path, use this setting

########## RUN START SETTING(!!! Follow below measure sequence !!!) ##########
SlaveAddr_0 = 0x2e # rdp180xp chip_id0
# logfile_path = None
logfile_path = f"/Users/rlee/Desktop/Ronlee/RDP180XP_electrical_test_results_{chip_num}.xlsx" # If you have an existing logfile_path, use this setting
start_time = time.time()

########### MEASUREMENT SETTING WITH CALIBRATION AT 50C -> CHAMBER TEMP SWEEP ##########
temp_lst = [50, -40, -20, 0, 20, 40, 60, 80, 90]
total_progress = 0
total_item_cal = 27
total_item_noncal = 18

for temp_progress, target_temp in enumerate(temp_lst):
    chip_num = f"{chip_num}_{target_temp}°C"
    print(f"CHIP_NUM : {chip_num}")
    chamber.set_target_temp(target_temp)
    chamber.start()
    testitem_progress = 0

    if target_temp == 50:
        cnt_item = 0
        while True:
            temp_curr = chamber.get_current_temp()   
            if (target_temp-0.1) <= temp_curr <= (target_temp+0.1):
                print("START")
                time.sleep(300) # 5 min for chamber settling
                record_test_info()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                i2c_all_register_test()
                total_progress -= 1
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                i2c_reset()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                i2c_all_power_up()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert Front Terminal(AT0 Measure)
                time.sleep(1) # Delay 1sec
                calibration_hvbgr_top_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_hvbgr_top_current()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_ldo33_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_adc_refh_top_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_adc_refl_top_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_top_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_top_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_top_code()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_top_code()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_sub_pixel_current()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                smu.gpib.write(":ROUTe:TERMinals REAR") # Convert REAR Terminal(AT1 Measure)
                time.sleep(1) # Delay 1sec
                print("CONVERT SMU TEMINAL!!!", end="\n\n")
                calibration_hvbgr_btm_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_hvbgr_btm_current()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_adc_refh_btm_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                calibration_adc_refl_btm_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_btm_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_btm_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_btm_code()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_btm_code()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_mipirx_bias_voltage()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_adc_external_sweep()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert Front Terminal(AT0 Measure)
                time.sleep(1) # Delay 1sec
                print("CONVERT SMU TEMINAL!!!", end="\n\n")
                measure_column_data_sweep()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                measure_bist_functions()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/1000) # 1ms
                read_calibration_values()
                update_gui_info(num_item=total_item_cal)
                time.sleep(1/100) # 10ms
                chip_num = chip_num.split("_")[0]
                break
    
    elif target_temp != 50:
        cnt_item = 0
        while True:
            temp_curr = chamber.get_current_temp()   
            if (target_temp-0.1) <= temp_curr <= (target_temp+0.1):
                print("START")
                time.sleep(300) # 5 min for chamber settling
                record_test_info()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000)
                i2c_all_register_test()
                total_progress -= 1
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000)
                i2c_reset()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000)
                write_calibration_values()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/100)
                i2c_all_power_up()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_atest_all() # AT0(Front) -> AT1(Rear)
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_top_voltage() #AT0(Front)
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_top_voltage() #AT0(Front)
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                smu.gpib.write(":ROUTe:TERMinals REAR") # Convert REAR Terminal(AT1 Measure)
                time.sleep(1) # Delay 1sec
                print("CONVERT SMU TEMINAL(REAR)!!!", end="\n\n")
                measure_tempsensor_left_btm_voltage() #AT1(Rear)
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_btm_voltage() #AT1(Rear)
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_top_code()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_top_code()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_left_btm_code()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_tempsensor_right_btm_code()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_adc_external_sweep() # AT0
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                smu.gpib.write(":ROUTe:TERMinals FRONt") # Convert FRONT Terminal(AT0 Measure)
                time.sleep(1) # Delay 1sec
                print("CONVERT SMU TEMINAL(FRONT)!!!", end="\n\n")
                measure_column_data_sweep() # AT0
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_sub_pixel_current() # AT0
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                measure_bist_functions()
                update_gui_info(num_item=total_item_noncal)
                time.sleep(1/1000) # 1ms
                chip_num = chip_num.split("_")[0]
                print("END")
                break

########### RUN CLOSE SETTING ##########
chamber.stop()
chamber.close()
chamber.power_off()
smu.reset()
smu.disable_output()
smu.close()
evb.close()
root.mainloop()

end_time = time.time()
print(f"{end_time - start_time:.5f} sec")